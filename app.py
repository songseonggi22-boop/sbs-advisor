"""
app.py — SBS아카데미 대전지점 상담 시스템 (Streamlit 웹 대시보드)
실행: streamlit run app.py
"""

from __future__ import annotations
from datetime import datetime
from pathlib import Path
import io
import base64
import json
import os
import threading
import time

import pandas as pd
import requests
import streamlit as st
import openpyxl
import google.generativeai as genai
import markdown as md_lib
from dotenv import load_dotenv

from design_engine import design, COURSE_ROLES, FIELDS
from excel_loader import load_courses as _excel_load_courses
from calculator import course_level_count

load_dotenv()
WP_URL          = os.getenv("WP_URL", "")
WP_USERNAME     = os.getenv("WP_USERNAME", "")
WP_APP_PASSWORD = os.getenv("WP_APP_PASSWORD", "")
PEXELS_API_KEY  = os.getenv("PEXELS_API_KEY", "")  # Pexels — 실사 이미지 검색용

# ── 경로 ─────────────────────────────────────────────────────────────────────
ROOT       = Path(__file__).parent
LOGS_DIR   = ROOT / "logs"
QUEUE_FILE = ROOT / "queue.json"
LOGS_DIR.mkdir(exist_ok=True)

# ── 페이지 설정 ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="SBS아카데미 대전지점 · 상담 시스템",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── 글로벌 CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── 사이드바 ── */
[data-testid="stSidebar"] { background:#111827; }
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] div { color:#e5e7eb !important; }
[data-testid="stSidebar"] h3  { color:#fff !important; font-weight:800; }
[data-testid="stSidebar"] hr  { border-color:#374151; }

/* ── 헤더 배너 ── */
.sbs-header {
    background: linear-gradient(120deg,#9b0010 0%,#e50012 55%,#ff3344 100%);
    padding:1.4rem 2rem; border-radius:14px; margin-bottom:1.2rem;
    box-shadow:0 4px 24px rgba(229,0,18,.30);
}
.sbs-header h1 { margin:0; color:#fff; font-size:1.75rem; font-weight:900; letter-spacing:-.5px; }
.sbs-header p  { margin:.25rem 0 0; color:rgba(255,255,255,.80); font-size:.9rem; }

/* ── 합계 카드 ── */
.total-card {
    background:linear-gradient(135deg,#9b0010,#e50012);
    border-radius:16px; padding:1.6rem 1.2rem;
    text-align:center; margin:0 0 1rem;
    box-shadow:0 6px 28px rgba(229,0,18,.35);
}
.total-label   { font-size:.85rem; color:rgba(255,255,255,.80); margin-bottom:.2rem; }
.total-amount  { font-size:2.8rem; font-weight:900; color:#fff; letter-spacing:-1.5px; line-height:1.1; }
.total-savings { font-size:.85rem; color:rgba(255,255,255,.75); margin-top:.4rem; }

/* ── 정보 카드 ── */
.info-card {
    background:#f9fafb; border:1.5px solid #e5e7eb;
    border-radius:12px; padding:1rem 1.2rem; margin-bottom:.8rem;
}
.info-card b { color:#e50012; }

/* ── 프리셋 버튼 ── */
div[data-testid="column"] button {
    width:100% !important; border-radius:10px !important;
    font-weight:700 !important; border:2px solid #e50012 !important;
    color:#e50012 !important; background:#fff5f5 !important;
    padding:.55rem .5rem !important; transition:all .15s !important;
}
div[data-testid="column"] button:hover {
    background:#e50012 !important; color:#fff !important;
}

/* ── 테이블 헤더 ── */
thead tr th {
    background:#e50012 !important; color:#fff !important;
    font-weight:700 !important;
}

/* ── 섹션 타이틀 ── */
.section-title {
    font-size:1.05rem; font-weight:800; color:#111827;
    border-left:4px solid #e50012; padding-left:.7rem;
    margin:1.2rem 0 .7rem;
}

/* ── 성공/경고 배너 ── */
.ok-banner {
    background:#f0fdf4; border:1.5px solid #16a34a;
    border-radius:10px; padding:.8rem 1.2rem; color:#15803d;
    font-weight:600; margin:.6rem 0;
}

/* ══ AI 설계 검색 ══════════════════════════════════════════ */
.ai-search-wrap {
    background:linear-gradient(135deg,#0f172a 0%,#1e1b4b 100%);
    border-radius:16px; padding:1.5rem 1.8rem; margin-bottom:1.4rem;
    border:1px solid #312e81;
    box-shadow:0 4px 24px rgba(99,102,241,.25);
}
.ai-search-wrap h3 {
    color:#a5b4fc; font-size:1rem; font-weight:800;
    margin:0 0 .15rem; letter-spacing:.3px;
}
.ai-search-wrap p { color:#94a3b8; font-size:.82rem; margin:0 0 .8rem; }

/* 설계 결과 카드 공통 */
.design-card {
    border-radius:14px; padding:1.3rem 1.4rem;
    height:100%; box-shadow:0 4px 18px rgba(0,0,0,.10);
}
.design-card-min {
    background:#fff; border:2px solid #e50012;
}
.design-card-max {
    background:linear-gradient(160deg,#1a1a2e 0%,#16213e 100%);
    border:2px solid #6366f1; color:#e2e8f0;
}
.dc-badge-min {
    display:inline-block; background:#e50012; color:#fff;
    font-size:.72rem; font-weight:800; padding:.25rem .65rem;
    border-radius:20px; margin-bottom:.55rem; letter-spacing:.5px;
}
.dc-badge-max {
    display:inline-block;
    background:linear-gradient(90deg,#6366f1,#8b5cf6);
    color:#fff; font-size:.72rem; font-weight:800;
    padding:.25rem .65rem; border-radius:20px;
    margin-bottom:.55rem; letter-spacing:.5px;
}
.dc-field-label { font-size:.8rem; opacity:.65; margin-bottom:.2rem; }
.dc-price-min { font-size:2rem; font-weight:900; color:#e50012; letter-spacing:-1px; line-height:1.1; }
.dc-price-max { font-size:2rem; font-weight:900; color:#a5b4fc; letter-spacing:-1px; line-height:1.1; }
.dc-course-tag {
    display:inline-block; background:#f3f4f6; color:#1f2937;
    border-radius:6px; padding:.18rem .55rem;
    font-size:.78rem; font-weight:600; margin:.18rem .18rem 0 0;
}
.dc-course-tag-dark {
    display:inline-block;
    background:rgba(99,102,241,.18); color:#c7d2fe;
    border-radius:6px; padding:.18rem .55rem;
    font-size:.78rem; font-weight:600; margin:.18rem .18rem 0 0;
    border:1px solid rgba(99,102,241,.35);
}
.dc-intent-min {
    background:#fff5f5; border-left:3px solid #e50012;
    border-radius:0 8px 8px 0; padding:.65rem .9rem;
    font-size:.82rem; color:#374151; margin:.8rem 0;
    line-height:1.6;
}
.dc-intent-max {
    background:rgba(99,102,241,.12);
    border-left:3px solid #6366f1;
    border-radius:0 8px 8px 0; padding:.65rem .9rem;
    font-size:.82rem; color:#cbd5e1; margin:.8rem 0;
    line-height:1.6;
}
.dc-relation-title { font-size:.78rem; font-weight:700; opacity:.65; margin:.7rem 0 .3rem; }
.dc-relation-item  { font-size:.78rem; line-height:1.55; margin-bottom:.35rem; }

/* 힌트 칩 */
.hint-chip {
    display:inline-block; background:rgba(165,180,252,.12);
    border:1px solid rgba(165,180,252,.3); color:#a5b4fc;
    border-radius:20px; padding:.2rem .75rem; font-size:.78rem;
    margin:.15rem; cursor:default;
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# 데이터 & 상수
# ══════════════════════════════════════════════════════════════════════════════

PRESETS: dict[str, dict] = {
    "⚡ 영상·모션 최소": {
        "courses": ["프리미어", "에펙", "모션에펙"],
        "desc": "영상 편집 핵심 | 3개 과정 | 약 3~4개월",
    },
    "🚀 영상·모션 최대": {
        "courses": ["발상과표현", "포토샵, 일러스트", "프리미어", "에펙", "모션에펙", "어드벤스 에펙", "시포디1~4", "모션포폴1~3"],
        "desc": "풀스펙 모션 포트폴리오 | 8개 과정 | 약 10~12개월",
    },
    "✏️ 그래픽 최소": {
        "courses": ["포토샵, 일러스트", "포토웍스, 디일러"],
        "desc": "그래픽 기초 입문 | 2개 과정 | 약 2개월",
    },
    "🏠 인테리어 최소": {
        "courses": ["실내건축이론", "캐드1~2", "스케치업1~2"],
        "desc": "인테리어 입문 | 3개 과정 | 약 3~4개월",
    },
    "🌐 웹디자인 풀": {
        "courses": ["포토웍스, 디일러", "웹1~3", "UIUX1~3", "웹포폴1~2"],
        "desc": "웹 퍼블리셔 완성 | 4개 과정 | 약 5~6개월",
    },
    "🤖 AI 크리에이터": {
        "courses": ["AI프롬프트1~2", "AI크리에이터-영상모션", "AI크리에이터-아트웍"],
        "desc": "AI 기반 크리에이터 | 3개 과정 | 최신 트렌드",
    },
}


# ══════════════════════════════════════════════════════════════════════════════
# 데이터 로드 (캐시)
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data(show_spinner="수강료 데이터 로딩 중...")
def load_courses() -> list[dict]:
    """엑셀 파일에서 수강료 데이터를 로드합니다 (크롤링 없음)."""
    return _excel_load_courses()


def course_lookup(courses: list[dict]) -> dict[str, int]:
    return {c["course"]: c["price"] for c in courses}


def build_category_map(courses: list[dict]) -> dict[str, list[str]]:
    cat_map: dict[str, list[str]] = {}
    for c in courses:
        dept = c["department"]
        if dept not in cat_map:
            cat_map[dept] = []
        cat_map[dept].append(c["course"])
    return cat_map


# ══════════════════════════════════════════════════════════════════════════════
# 계산 유틸
# ══════════════════════════════════════════════════════════════════════════════

def fmt_won(n: int) -> str:
    return f"{n:,}원"


def build_df(selected_names: list[str], price_map: dict[str, int], discount_rate: float) -> pd.DataFrame:
    rows = []
    for name in selected_names:
        unit  = price_map.get(name, 0)
        lvls  = course_level_count(name)
        price = unit * lvls
        disc_price = round(price * (1 - discount_rate / 100))
        level_str = f"{lvls}단계" if lvls > 1 else "1"
        rows.append({
            "과정명":   name,
            "단가":     fmt_won(unit),
            "단계":     level_str,
            "정가":     fmt_won(price),
            "할인율":   f"{discount_rate:.0f}%",
            "할인가":   fmt_won(disc_price),
            "절약":     fmt_won(price - disc_price),
            "_unit":    unit,
            "_lvls":    lvls,
            "_price":   price,
            "_final":   disc_price,
        })
    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
# 문서 생성 (MD / HTML)
# ══════════════════════════════════════════════════════════════════════════════

def build_md(name, contact, field, consultant, memo, df, total, subtotal, savings, disc_rate, today) -> str:
    lines = [
        "# SBS아카데미 대전지점 상담일지", "",
        f"> 작성일시: {today}  ",
        f"> 담당 상담사: {consultant}", "",
        "---", "",
        "## 상담생 정보", "",
        f"- **이름**: {name}",
        f"- **연락처**: {contact}",
        f"- **희망 분야**: {field}", "",
        "## 수강료 내역", "",
        "| 과정명 | 단가 | 단계 | 정가 | 할인율 | 할인가 | 절약 |",
        "|---|---|---|---|---|---|---|",
    ]
    for _, r in df.iterrows():
        lines.append(f"| {r['과정명']} | {r['단가']} | {r['단계']} | {r['정가']} | {r['할인율']} | {r['할인가']} | {r['절약']} |")
    lines += [
        "",
        f"- **소계**: {fmt_won(subtotal)}",
        f"- **할인율**: {disc_rate:.0f}%",
        f"- **최종 납부금액**: **{fmt_won(total)}**",
        f"- **총 절약금액**: {fmt_won(savings)}", "",
        "---", "",
        "## 상담 메모", "",
        memo.strip() or "*(메모 없음)*", "",
        "---", "",
        "## 팀장 확인 및 코멘트", "",
        "| 항목 | 내용 |",
        "|---|---|",
        "| 검토 여부 | ☐ 검토 완료 |",
        "| 등록 여부 | ☐ 등록 / ☐ 미등록 / ☐ 보류 |",
        "| 팀장 서명 | |",
        "| 코멘트 | |", "",
        "> *이 문서는 SBS아카데미 대전지점 내부용 상담 기록입니다.*",
    ]
    return "\n".join(lines)


def build_html(name, contact, field, consultant, memo, df, total, subtotal, savings, disc_rate, today) -> str:
    rows_html = "".join(
        f"<tr><td>{r['과정명']}</td><td>{r['단가']}</td><td style='text-align:center'>{r['단계']}</td>"
        f"<td>{r['정가']}</td>"
        f"<td style='color:#e50012'>{r['할인율']}</td>"
        f"<td style='font-weight:700'>{r['할인가']}</td>"
        f"<td style='color:#16a34a'>{r['절약']}</td></tr>"
        for _, r in df.iterrows()
    )
    memo_html = memo.strip().replace("\n", "<br>") or "<em>(메모 없음)</em>"

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<title>SBS아카데미 상담일지 – {name}</title>
<style>
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ font-family:'Malgun Gothic',Arial,sans-serif; color:#1f2937; padding:2.5rem; font-size:14px; }}
  .header {{ background:linear-gradient(120deg,#9b0010,#e50012); color:#fff; padding:1.5rem 2rem; border-radius:10px; margin-bottom:1.5rem; }}
  .header h1 {{ font-size:1.5rem; font-weight:900; }}
  .header p  {{ opacity:.85; margin-top:.3rem; font-size:.88rem; }}
  h2 {{ font-size:1rem; font-weight:800; color:#e50012; border-left:4px solid #e50012;
        padding-left:.6rem; margin:1.4rem 0 .7rem; }}
  .info-grid {{ display:grid; grid-template-columns:1fr 1fr; gap:.5rem .5rem; margin-bottom:.5rem; }}
  .info-item {{ background:#f9fafb; border:1px solid #e5e7eb; border-radius:8px; padding:.6rem .9rem; }}
  .info-item span {{ font-size:.78rem; color:#6b7280; display:block; }}
  .info-item b   {{ font-size:.95rem; color:#111827; }}
  table {{ width:100%; border-collapse:collapse; margin-bottom:1rem; }}
  th {{ background:#e50012; color:#fff; padding:.6rem .8rem; text-align:left; font-size:.85rem; }}
  td {{ padding:.55rem .8rem; border-bottom:1px solid #f3f4f6; font-size:.88rem; }}
  tr:nth-child(even) td {{ background:#fef2f2; }}
  .summary-box {{ background:linear-gradient(135deg,#9b0010,#e50012); color:#fff;
                  border-radius:12px; padding:1.2rem 1.5rem; margin:1rem 0; text-align:center; }}
  .summary-box .lbl {{ font-size:.82rem; opacity:.8; }}
  .summary-box .amt {{ font-size:2.2rem; font-weight:900; letter-spacing:-1px; }}
  .summary-box .sav {{ font-size:.82rem; opacity:.8; margin-top:.2rem; }}
  .memo-box {{ background:#f9fafb; border:1.5px solid #e5e7eb; border-radius:10px; padding:1rem; margin-bottom:1rem; }}
  .check-table td {{ padding:.5rem .8rem; }}
  .footer {{ color:#9ca3af; font-size:.78rem; text-align:center; margin-top:2rem; border-top:1px solid #e5e7eb; padding-top:1rem; }}
  @media print {{
    body {{ padding:1rem; }}
    .no-print {{ display:none; }}
  }}
</style>
</head>
<body>
<div class="header">
  <h1>🎓 SBS아카데미 대전지점 상담일지</h1>
  <p>작성일시: {today} &nbsp;|&nbsp; 담당 상담사: {consultant}</p>
</div>

<h2>상담생 정보</h2>
<div class="info-grid">
  <div class="info-item"><span>이름</span><b>{name}</b></div>
  <div class="info-item"><span>연락처</span><b>{contact}</b></div>
  <div class="info-item"><span>희망 분야</span><b>{field}</b></div>
  <div class="info-item"><span>담당 상담사</span><b>{consultant}</b></div>
</div>

<h2>수강료 내역</h2>
<table>
  <thead><tr><th>과정명</th><th>단가</th><th>단계</th><th>정가</th><th>할인율</th><th>할인가</th><th>절약</th></tr></thead>
  <tbody>{rows_html}</tbody>
</table>

<div class="summary-box">
  <div class="lbl">최종 납부금액</div>
  <div class="amt">{fmt_won(total)}</div>
  <div class="sav">할인율 {disc_rate:.0f}% 적용 · 총 {fmt_won(savings)} 절약</div>
</div>

<h2>상담 메모</h2>
<div class="memo-box">{memo_html}</div>

<h2>팀장 확인 및 코멘트</h2>
<table class="check-table">
  <tr><td style="width:140px;color:#6b7280">검토 여부</td><td>☐ 검토 완료</td></tr>
  <tr><td style="color:#6b7280">등록 여부</td><td>☐ 등록 &nbsp; ☐ 미등록 &nbsp; ☐ 보류</td></tr>
  <tr><td style="color:#6b7280">팀장 서명</td><td>&nbsp;</td></tr>
  <tr><td style="color:#6b7280">코멘트</td><td>&nbsp;</td></tr>
</table>

<p class="footer">이 문서는 SBS아카데미 대전지점 내부용 상담 기록입니다.</p>
</body>
</html>"""


# ══════════════════════════════════════════════════════════════════════════════
# 견적서 셀 위치 맵 — 양식이 바뀌면 이 블록만 수정하세요
# ══════════════════════════════════════════════════════════════════════════════
#
# ▸ 고정 셀: 행/열이 항상 같은 셀의 전체 주소 (예: "D5")
# ▸ 동적 셀: 과정 수에 따라 행이 달라지므로 열 문자와 기준 행 상수로 관리
#            실제 주소는 build_excel 내부의 _make_cell_map()이 조합합니다.
#
CELL_MAP: dict[str, str | int] = {
    # ── 과정 데이터 (고정 24행: D7~D30 = 과정명, E7~E30 = 등록비) ─────────────
    "과정_시작행":  7,      # 첫 번째 과정 데이터 행
    "과정_종료행":  30,     # 마지막 과정 데이터 행
    "과정명_열":    "D",    # 과정명 입력 열
    "등록비_열":    "E",    # 수강료 입력 열
    # ── 합계 ──────────────────────────────────────────────────────────────────
    "합계셀":       "E31",  # 총 등록비 (SUM 수식)
    # ── 할인 섹션 (35~39행 고정, A=명분, C=율, D=금액, E=잔액) ─────────────────
    "할인_시작행":  35,
    "명분_열":      "A",    # A35:B35 병합 → A열에 입력
    "할인율_열":    "C",
    "할인금액_열":  "D",
    "잔액_열":      "E",
    # ── 최종 합계 (병합셀 → 첫 번째 셀에 입력) ──────────────────────────────
    "납부액셀":     "C40",  # 총 납부해야할 등록비 (C40:E40 병합)
    "온라인셀":     "C41",  # 따따아 온라인 수강비 990,000 (변경 안 함)
    "최종셀":       "C42",  # 총 납부할 수강료 (C42:E42 병합)
}

# ══════════════════════════════════════════════════════════════════════════════
# 엑셀 견적서 생성 헬퍼
# ══════════════════════════════════════════════════════════════════════════════



# ══════════════════════════════════════════════════════════════════════════════
# 엑셀 견적서 생성 (견적서_양식.xlsx 템플릿 기반 — 서식 100% 유지)
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(
    name: str, contact: str, field: str, consultant: str, memo: str,
    df: "pd.DataFrame", total: int, subtotal: int, savings: int,
    disc_rate: float, today: str,
    disc_rate1: float = 0.0, disc_reason1: str = "",
    disc_rate2: float = 0.0, disc_reason2: str = "",
    disc_rate3: float = 0.0, disc_reason3: str = "",
    disc_rate4: float = 0.0, disc_reason4: str = "",
    disc_rate5: float = 0.0, disc_reason5: str = "",
) -> bytes:
    """패키지 견적서(고정).xlsx 템플릿을 로드해 CELL_MAP 기반으로 데이터를 기입합니다.
    병합 셀·색상·폰트·테두리 등 서식을 100% 유지합니다."""

    TEMPLATE = ROOT / "패키지 견적서(고정).xlsx"
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb.worksheets[0]   # 시트명: 견적서

    cm       = CELL_MAP
    start    = int(cm["과정_시작행"])   # 7
    end      = int(cm["과정_종료행"])   # 30
    col_name = cm["과정명_열"]          # D
    col_fee  = cm["등록비_열"]          # E

    # ── 1. 과정 데이터 채우기 (최대 24과목) ──────────────────────────────────
    for i, (_, row) in enumerate(df.iterrows()):
        r = start + i
        if r > end:
            break
        ws[f"{col_name}{r}"] = row["과정명"]
        ws[f"{col_fee}{r}"]  = row["_price"]

    # ── 2. 합계 수식 (E31 = SUM(E7:E30)) ─────────────────────────────────────
    ws[cm["합계셀"]] = f"=SUM({col_fee}{start}:{col_fee}{end})"

    # ── 3. 단계별 차감(Compound Discount) 수식 (35~39행) ─────────────────────
    _DEFAULT_REASONS = [
        "8개 과목이상 수강시 패키지 할인",
        "따따아 과정 우선 등록",
        "첫번째 첫 결제 패키지 할인",
        "원장 결재권한 할인",
        "기타 할인",
    ]
    disc_rates   = [disc_rate1,   disc_rate2,   disc_rate3,   disc_rate4,   disc_rate5]
    disc_reasons = [disc_reason1, disc_reason2, disc_reason3, disc_reason4, disc_reason5]
    hr      = int(cm["할인_시작행"])   # 35
    col_a   = cm["명분_열"]             # A (A35:B35 병합 → A에 입력)
    col_c   = cm["할인율_열"]           # C
    col_d   = cm["할인금액_열"]         # D
    col_e   = cm["잔액_열"]             # E
    prev_e  = cm["합계셀"]              # E31 (첫 번째 단계의 기준)

    for i, (rate, reason) in enumerate(zip(disc_rates, disc_reasons)):
        r        = hr + i   # 35, 36, 37, 38, 39
        rate_dec = round(rate / 100, 4)
        ws[f"{col_a}{r}"] = reason or _DEFAULT_REASONS[i]
        ws[f"{col_c}{r}"] = rate_dec
        # 할인금액 = 이전 잔액 × 이 단계 할인율
        ws[f"{col_d}{r}"] = f"={prev_e}*{col_c}{r}"
        # 잔액 = 이전 잔액 × (1 - 이 단계 할인율)
        ws[f"{col_e}{r}"] = f"={prev_e}*(1-{col_c}{r})"
        prev_e = f"{col_e}{r}"

    # ── 4. 최종 합계 수식 ─────────────────────────────────────────────────────
    last_e_row = hr + len(disc_rates) - 1        # 39
    last_e     = f"{col_e}{last_e_row}"           # E39
    ws[cm["납부액셀"]] = f"={last_e}"             # C40 = E39
    # C41(따따아 온라인 수강비 990,000)은 템플릿 값 유지 — 변경 안 함
    ws[cm["최종셀"]] = f"={cm['납부액셀']}+{cm['온라인셀']}"  # C42 = C40 + C41

    # ── 5. bytes 반환 ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# 워드프레스 발행 유틸
# ══════════════════════════════════════════════════════════════════════════════

def extract_title(md_text: str) -> str:
    """마크다운에서 첫 번째 H1 제목을 추출합니다."""
    for line in md_text.splitlines():
        stripped = line.strip()
        if stripped.startswith("# "):
            return stripped[2:].strip()
    return "SBS아카데미 블로그 포스트"


def extract_image_prompt(md_text: str) -> tuple[str, str]:
    """[IMAGE_KEYWORD]: 줄을 추출하고 나머지 정제된 텍스트를 반환합니다. (구버전 [IMAGE_PROMPT]: 도 호환)"""
    lines = md_text.splitlines()
    image_keyword = ""
    clean_lines = []
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("[IMAGE_KEYWORD]:"):
            image_keyword = stripped[len("[IMAGE_KEYWORD]:"):].strip()
        elif stripped.startswith("[IMAGE_PROMPT]:"):
            image_keyword = stripped[len("[IMAGE_PROMPT]:"):].strip()
        else:
            clean_lines.append(line)
    return "\n".join(clean_lines).strip(), image_keyword


def fetch_pexels_image(keyword: str, pexels_key: str) -> dict:
    """Pexels API로 키워드에 맞는 고해상도 실사 이미지 정보를 반환합니다."""
    resp = requests.get(
        "https://api.pexels.com/v1/search",
        headers={"Authorization": pexels_key},
        params={"query": keyword, "per_page": 1, "orientation": "landscape"},
        timeout=15,
    )
    resp.raise_for_status()
    photos = resp.json().get("photos", [])
    if not photos:
        return {}
    photo = photos[0]
    return {
        "url":          photo["src"]["large2x"],
        "photographer": photo.get("photographer", "Pexels"),
        "page_url":     photo.get("url", "https://www.pexels.com"),
    }


def insert_pexels_image_html(html_content: str, img_info: dict, alt: str) -> str:
    """서론 첫 단락 바로 아래에 Pexels 이미지와 출처를 삽입합니다."""
    url          = img_info.get("url", "")
    photographer = img_info.get("photographer", "Pexels")
    page_url     = img_info.get("page_url", "https://www.pexels.com")
    if not url:
        return html_content
    img_tag = (
        f'<figure style="text-align:center;margin:1.5rem 0;">'
        f'<img src="{url}" alt="{alt}" style="max-width:100%;height:auto;border-radius:8px;">'
        f'<figcaption style="color:#6b7280;font-size:.8rem;margin-top:.4rem;">'
        f'Photo by <a href="{page_url}" target="_blank" rel="noopener">{photographer}</a>'
        f' on <a href="https://www.pexels.com" target="_blank" rel="noopener">Pexels</a>'
        f'</figcaption>'
        f'</figure>'
    )
    idx = html_content.find("</p>")
    if idx != -1:
        return html_content[:idx + 4] + "\n" + img_tag + "\n" + html_content[idx + 4:]
    return img_tag + "\n" + html_content


def insert_image_html(html_content: str, img_url: str, alt: str) -> str:
    """(레거시 호환) insert_pexels_image_html 래퍼."""
    return insert_pexels_image_html(html_content, {"url": img_url}, alt)


# ── 큐 영속성 ──────────────────────────────────────────────────────────────────

def load_queue() -> list:
    """queue.json에서 키워드 큐를 로드합니다."""
    try:
        if QUEUE_FILE.exists():
            return json.loads(QUEUE_FILE.read_text(encoding="utf-8"))
    except Exception:
        pass
    return []


def save_queue(queue: list) -> None:
    """큐를 queue.json에 저장합니다."""
    QUEUE_FILE.write_text(json.dumps(queue, ensure_ascii=False, indent=2), encoding="utf-8")


# ── 백그라운드 발행 스레드 ─────────────────────────────────────────────────────

_BG_STATE: dict = {"running": False, "stop": False, "thread": None}


def _bg_worker(interval_secs: int, gemini_key: str, pexels_key: str,
               wp_status: str, include_img: bool) -> None:
    """백그라운드 스레드: 큐에서 키워드를 순서대로 꺼내 발행합니다."""
    _BG_STATE["running"] = True
    _BG_STATE["stop"]    = False
    try:
        while not _BG_STATE["stop"]:
            queue   = load_queue()
            pending = [(i, item) for i, item in enumerate(queue) if item["status"] == "pending"]
            if not pending:
                break
            idx, item = pending[0]
            kw = item["keyword"]
            queue[idx].update({"status": "processing",
                                "started_at": datetime.now().strftime("%Y-%m-%d %H:%M")})
            save_queue(queue)
            try:
                # 1. 원고 생성
                genai.configure(api_key=gemini_key)
                model = genai.GenerativeModel(
                    model_name="gemini-2.5-flash",
                    system_instruction=BLOG_DEFAULT_PROMPT,
                )
                clean_text, img_keyword = extract_image_prompt(
                    model.generate_content(f"결과를 작성할 메인 키워드: {kw}").text
                )
                title = extract_title(clean_text)
                # 2. Pexels 이미지
                img_info   = {}
                img_status = "—"
                if include_img and pexels_key and img_keyword:
                    try:
                        img_info   = fetch_pexels_image(img_keyword, pexels_key)
                        img_status = "✅ 완료" if img_info else "⚠️ 결과없음"
                    except Exception:
                        img_status = "❌ 실패"
                # 3. HTML + 발행
                html = md_lib.markdown(clean_text, extensions=["tables", "fenced_code"])
                if img_info.get("url"):
                    html = insert_pexels_image_html(html, img_info, kw)
                wp_result = post_to_wordpress(title, html, status=wp_status)
                queue = load_queue()
                queue[idx].update({
                    "status":       "done",
                    "img_status":   img_status,
                    "post_id":      wp_result.get("id"),
                    "post_link":    wp_result.get("link", ""),
                    "completed_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "error":        "",
                })
                save_queue(queue)
            except Exception as e:
                queue = load_queue()
                queue[idx].update({"status": "failed", "error": str(e)[:120]})
                save_queue(queue)
            # 다음 항목 대기
            if not _BG_STATE["stop"]:
                if any(i["status"] == "pending" for i in load_queue()):
                    time.sleep(interval_secs)
    finally:
        _BG_STATE["running"] = False


def post_to_wordpress(title: str, content: str, status: str = "publish") -> dict:
    """워드프레스 REST API로 포스트를 발행합니다. (Basic Auth + JSON)"""
    endpoint = f"{WP_URL.rstrip('/')}/wp-json/wp/v2/posts"
    token = base64.b64encode(
        f"{WP_USERNAME}:{WP_APP_PASSWORD}".encode("utf-8")
    ).decode("utf-8")
    headers = {
        "Authorization": f"Basic {token}",
        "Content-Type": "application/json",
    }
    resp = requests.post(
        endpoint,
        headers=headers,
        json={"title": title, "content": content, "status": status},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


# ══════════════════════════════════════════════════════════════════════════════
# 블로그 자동화 기본 시스템 프롬프트
# ══════════════════════════════════════════════════════════════════════════════

BLOG_DEFAULT_PROMPT = """\
# Role
너는 네이버 SEO + 구글 SEO에 동시에 최적화된 한국어 블로그 콘텐츠 라이터이자 편집자다. 특히 네이버 블로그 모바일 화면에서 가독성이 좋아지도록 "짧은 문단 + 잦은 줄바꿈"을 기본 규칙으로 원고를 작성한다. 후기형(리뷰 톤)으로 자연스럽게 쓰되, 구조(H1~H4)와 맞춤법·띄어쓰기·오탈자 품질을 최상으로 유지한다.

# Context
사용자가 메인키워드 1개를 주면, 네이버 SEO(가독성/체류시간/목차/후기형 문장/키워드 분산) + 구글 SEO(명확한 헤딩 계층, 스캐너블 구성, Q&A)를 만족하는 글을 작성한다.
- 이모티콘/이모지 완전 금지(숫자 이모지 포함 모든 유니코드 이모지 금지)
- 본문에 마크다운 표 최소 1개 포함
- 결과물은 바로 복사-붙여넣기 가능한 마크다운으로만 출력
- 홍보 대상: SBS아카데미컴퓨터학원 대전점
- 문단 규칙: 한 문단 1~3문장, 220자 이내 준수

# Task
- 구성: 제목(H1), 도입, 목차, 본문(H2/H3/H4), 표, Q&A(5개), 마무리, CTA(1개)
- CTA는 마지막 줄에 "SBS아카데미컴퓨터학원 대전점" 문의 유도로 고정

# Image Keyword
원고 작성이 끝나면 반드시 본문 맨 마지막 줄에 아래 형식으로 Pexels 이미지 검색 키워드 1개를 추가한다.
다른 내용과 반드시 한 줄 띄우고 출력한다:
[IMAGE_KEYWORD]: (글 전체 주제와 가장 잘 어울리는 영어 이미지 검색 단어 또는 짧은 구문 1개)
예시: [IMAGE_KEYWORD]: graphic design student studying
"""

# ══════════════════════════════════════════════════════════════════════════════
# Session State 초기화
# ══════════════════════════════════════════════════════════════════════════════

if "selected_courses" not in st.session_state:
    st.session_state.selected_courses: list[str] = []
if "preset_trigger" not in st.session_state:
    st.session_state.preset_trigger: str | None = None
if "ai_result" not in st.session_state:
    st.session_state.ai_result: dict | None = None
if "ai_query" not in st.session_state:
    st.session_state.ai_query: str = ""
if "ai_design_memo" not in st.session_state:
    st.session_state.ai_design_memo: str = ""
if "mentor_query" not in st.session_state:
    st.session_state.mentor_query: str = ""
if "mentor_answer" not in st.session_state:
    st.session_state.mentor_answer: str = ""
if "blog_keyword" not in st.session_state:
    st.session_state.blog_keyword: str = ""
if "blog_result" not in st.session_state:
    st.session_state.blog_result: str = ""
if "blog_system_prompt" not in st.session_state:
    st.session_state.blog_system_prompt: str = ""  # 첫 렌더 시 BLOG_DEFAULT_PROMPT로 채움
if "image_keyword" not in st.session_state:
    st.session_state.image_keyword: str = ""
if "pexels_img_info" not in st.session_state:
    st.session_state.pexels_img_info: dict = {}
if "generated_img_url" not in st.session_state:
    st.session_state.generated_img_url: str = ""
if "dashboard_results" not in st.session_state:
    st.session_state.dashboard_results: list = []


# ══════════════════════════════════════════════════════════════════════════════
# 사이드바
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("### 🎓 SBS아카데미 대전지점")
    menu = st.radio(
        "메뉴",
        ["📄 패키지 견적서", "✍️ 블로그 자동화", "📅 예약 대시보드"],
        label_visibility="collapsed",
    )
    st.markdown("---")

    st.markdown("**👤 상담생 정보**")
    s_name       = st.text_input("성함",     placeholder="홍길동")
    s_contact    = st.text_input("연락처",   placeholder="010-0000-0000")
    s_field      = st.text_input("희망 분야", placeholder="영상편집 / UI디자인 …")
    s_consultant = st.text_input("담당 상담사", value="상담사", placeholder="김상담")

    st.markdown("---")
    st.markdown("**💸 할인 설정**")

    # ── 단계별 할인 명분 + 할인율 입력 ──────────────────────────────────────
    st.markdown("**📝 단계별 할인 명분 & 할인율**")
    _DEFAULT_DISC_REASONS = [
        "8개 과목이상 수강시 패키지 할인",
        "따따아 과정 우선 등록",
        "첫번째 첫 결제 패키지 할인",
        "원장 결재권한 할인",
        "기타 할인",
    ]
    _disc_rates_list   = []
    _disc_reasons_list = []
    for _i, _def in enumerate(_DEFAULT_DISC_REASONS):
        _ca, _cb = st.columns([3, 1])
        with _ca:
            _disc_reasons_list.append(
                st.text_input(f"명분 {['①','②','③','④','⑤'][_i]}", value=_def, key=f"dr_reason{_i+1}")
            )
        with _cb:
            _disc_rates_list.append(
                st.number_input(f"율{['①','②','③','④','⑤'][_i]} %", min_value=0, max_value=40,
                                value=0, step=5, key=f"dr{_i+1}")
            )
    disc_rate1, disc_rate2, disc_rate3, disc_rate4, disc_rate5 = _disc_rates_list
    disc_reason1, disc_reason2, disc_reason3, disc_reason4, disc_reason5 = _disc_reasons_list

    # 단계별 차감(Compound Discount) 최종 할인율 자동 계산
    _rs = [r / 100 for r in _disc_rates_list]
    s_discount = round((1 - (1-_rs[0])*(1-_rs[1])*(1-_rs[2])*(1-_rs[3])*(1-_rs[4])) * 100, 2)
    if any(_disc_rates_list):
        st.info(
            f"복합 할인율: **{s_discount:.1f}%**  \n"
            f"({' → '.join(f'{r}%' for r in _disc_rates_list if r)} 순차 적용)",
        )

    # ── 참고: 과목수 기반 자동 계산 ──────────────────────────────────────────
    with st.expander("💡 과목수 기준 자동 계산 참고"):
        disc_type = st.radio(
            "할인 유형",
            ["일반 (과목수 자동)", "페북/구디비/종강생"],
            index=0,
            help="페북/구디비/종강생: 1~3과목 30%, 4과목 이상 40%",
        )
        extra_disc = st.select_slider(
            "추가 할인",
            options=[0, 5, 10],
            value=0,
            format_func=lambda x: f"+{x}%",
            help="10% 이내 자율 추가 할인 (5% 단위)",
        )
        _n_ref = sum(
            course_level_count(c)
            for c in st.session_state.get("selected_courses", [])
        )
        if disc_type == "페북/구디비/종강생":
            _sug = 40 if _n_ref >= 4 else 30
        else:
            if _n_ref <= 1:   _sug = 0
            elif _n_ref <= 3: _sug = 10
            elif _n_ref <= 5: _sug = 15
            elif _n_ref <= 7: _sug = 20
            else:             _sug = 25
        _sug = min(_sug + extra_disc, 40)
        st.info(f"총 **{_n_ref}단계** 기준 추천 할인율: **{_sug}%**  \n"
                f"(위 슬라이더에 직접 입력해 적용하세요)")

    # ── 추가 금액 할인 ────────────────────────────────────────────────────────
    review_on = st.checkbox("리뷰 할인 (최대 20,000원)", help="네이버·구글·카카오맵 리뷰 작성 시")
    review_won = st.number_input("리뷰 할인 금액(원)", 0, 20000, 10000, step=1000,
                                  label_visibility="collapsed") if review_on else 0
    ddaza_on = st.checkbox("따즈아 등록 할인 (최대 100,000원)",
                            help="오프라인 등록 할인율 적용 후 150만 이상인 경우 가능")
    ddaza_won = st.number_input("따즈아 할인 금액(원)", 0, 100000, 0, step=10000,
                                 label_visibility="collapsed") if ddaza_on else 0

    st.markdown("---")
    st.markdown("**📝 상담 메모**")
    _memo_default = st.session_state.get("ai_design_memo", "")
    s_memo = st.text_area("메모", value=_memo_default, height=160,
                           placeholder="수강 목표, 희망 수업 시간, 특이사항 등을 입력하세요.")

    st.markdown("---")
    st.markdown("**🤖 AI 멘토링 설정**")
    gemini_api_key = st.text_input(
        "Gemini API Key",
        type="password",
        placeholder="AIza...",
        help="Google AI Studio에서 발급한 Gemini API Key를 입력하세요.",
    )
    pexels_api_key = st.text_input(
        "Pexels API Key (이미지 검색용)",
        type="password",
        placeholder="BzAz...",
        value=PEXELS_API_KEY,
        help="pexels.com/api 에서 무료 발급. 실사 이미지 자동 삽입에 사용됩니다.",
    )

    st.markdown("---")
    if st.button("🔄 수강료 데이터 새로고침", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# 메인 화면
# ══════════════════════════════════════════════════════════════════════════════

# 헤더
st.markdown("""
<div class="sbs-header">
  <h1>🎓 SBS아카데미 대전지점 · 교육과정 상담 시스템</h1>
  <p>수강료 계산 · 커리큘럼 설계 · 상담일지 자동 생성</p>
</div>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# 블로그 자동화 탭
# ══════════════════════════════════════════════════════════════════════════════

if menu == "✍️ 블로그 자동화":
    st.markdown("""
<div style="background:linear-gradient(120deg,#064e3b 0%,#065f46 60%,#047857 100%);
            padding:1.2rem 2rem;border-radius:14px;margin-bottom:1.4rem;
            box-shadow:0 4px 24px rgba(5,150,105,.30);">
  <h2 style="margin:0;color:#fff;font-size:1.4rem;font-weight:900;">블로그 자동화 · SEO 원고 생성</h2>
  <p style="margin:.25rem 0 0;color:rgba(255,255,255,.80);font-size:.9rem;">
    Gemini 1.5 Flash 기반 · 네이버/구글 SEO 최적화 · SBS아카데미컴퓨터학원 대전점
  </p>
</div>
""", unsafe_allow_html=True)

    # 시스템 프롬프트가 초기화되지 않은 경우 기본값 세팅
    if not st.session_state.blog_system_prompt:
        st.session_state.blog_system_prompt = BLOG_DEFAULT_PROMPT

    with st.expander("⚙️ 시스템 프롬프트 설정 (고급)", expanded=False):
        st.caption("아래 내용을 직접 수정하면 생성 스타일과 지침을 변경할 수 있습니다.")
        st.text_area(
            "시스템 프롬프트",
            key="blog_system_prompt",
            height=380,
            label_visibility="collapsed",
        )
        if st.button("기본값으로 초기화", key="reset_prompt"):
            st.session_state.blog_system_prompt = BLOG_DEFAULT_PROMPT
            st.rerun()

    st.markdown("**메인 키워드 입력**")
    blog_keyword = st.text_input(
        "메인 키워드",
        value=st.session_state.blog_keyword,
        placeholder="예: SBS아카데미 대전 영상편집 과정",
        label_visibility="collapsed",
    )

    if st.button("✍️ 원고 생성", type="primary", use_container_width=True):
        if not gemini_api_key:
            st.warning("사이드바 하단에서 Gemini API Key를 입력해 주세요.")
        elif not blog_keyword.strip():
            st.warning("메인 키워드를 입력해 주세요.")
        else:
            st.session_state.blog_keyword = blog_keyword.strip()
            with st.spinner("블로그 원고를 생성 중입니다... (30초~1분 소요)"):
                try:
                    genai.configure(api_key=gemini_api_key)
                    model = genai.GenerativeModel(
                        model_name="gemini-2.5-flash",
                        system_instruction=st.session_state.blog_system_prompt,
                    )
                    response = model.generate_content(
                        f"결과를 작성할 메인 키워드: {blog_keyword.strip()}"
                    )
                    clean_text, img_keyword = extract_image_prompt(response.text)
                    st.session_state.blog_result       = clean_text
                    st.session_state.image_keyword     = img_keyword
                    st.session_state.generated_img_url = ""
                    st.session_state.pexels_img_info   = {}
                except Exception as e:
                    err_str = str(e)
                    if "429" in err_str or "quota" in err_str.lower():
                        st.error("API 사용량이 초과되었습니다. 잠시 후 다시 시도해 주세요.")
                    else:
                        st.error(f"오류가 발생했습니다: {e}")

    if st.session_state.blog_result:
        st.markdown("---")
        col_a, col_b = st.columns([1, 1])
        with col_a:
            st.markdown("**복사용 원고** (전체 선택 후 Ctrl+C)")
        with col_b:
            if st.button("✕ 원고 초기화", key="clear_blog"):
                st.session_state.blog_result       = ""
                st.session_state.image_keyword     = ""
                st.session_state.generated_img_url = ""
                st.session_state.pexels_img_info   = {}
                st.rerun()

        st.text_area(
            "원고",
            value=st.session_state.blog_result,
            height=500,
            label_visibility="collapsed",
        )

        # ── Pexels 실사 이미지 검색 ───────────────────────────────────────────
        st.markdown("---")
        st.markdown("**📷 Pexels 실사 이미지 자동 삽입**")

        if st.session_state.image_keyword:
            st.caption(f"AI 추천 검색어: `{st.session_state.image_keyword}`")
        else:
            st.caption("이미지 검색 키워드가 없습니다. (원고 재생성 시 자동 추출됩니다)")

        img_col1, img_col2 = st.columns([3, 1])
        with img_col1:
            custom_img_kw = st.text_input(
                "Pexels 검색 키워드 (직접 수정 가능)",
                value=st.session_state.image_keyword,
                key="custom_img_kw",
                label_visibility="collapsed",
            )
        with img_col2:
            search_img_btn = st.button("🔍 이미지 검색", use_container_width=True, key="search_img")

        if search_img_btn:
            if not pexels_api_key:
                st.warning("사이드바에서 Pexels API Key를 입력해 주세요.")
            elif not custom_img_kw.strip():
                st.warning("이미지 검색 키워드를 입력해 주세요.")
            else:
                with st.spinner("Pexels에서 이미지 검색 중..."):
                    try:
                        img_info = fetch_pexels_image(custom_img_kw.strip(), pexels_api_key)
                        if img_info:
                            st.session_state.pexels_img_info   = img_info
                            st.session_state.generated_img_url = img_info["url"]
                            st.success("이미지 검색 완료!")
                        else:
                            st.warning("검색 결과가 없습니다. 다른 키워드를 시도해 주세요.")
                    except Exception as e:
                        st.error(f"이미지 검색 실패: {e}")

        if st.session_state.generated_img_url:
            st.image(st.session_state.generated_img_url,
                     caption="Pexels 이미지 미리보기", use_container_width=True)
            if st.session_state.pexels_img_info.get("photographer"):
                st.caption(f"Photo by {st.session_state.pexels_img_info['photographer']} on Pexels")

        # ── 미리보기 ───────────────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("**미리보기**")
        st.markdown(st.session_state.blog_result)

        # ── 워드프레스 발행 ───────────────────────────────────────────────────
        st.markdown("---")
        if WP_URL and WP_USERNAME and WP_APP_PASSWORD:
            wp_col1, wp_col2 = st.columns(2)
            with wp_col1:
                if st.button("🚀 워드프레스에 바로 발행하기", type="primary",
                             use_container_width=True, key="wp_publish"):
                    with st.spinner("워드프레스에 발행 중..."):
                        try:
                            post_title   = extract_title(st.session_state.blog_result)
                            html_content = md_lib.markdown(
                                st.session_state.blog_result,
                                extensions=["tables", "fenced_code"],
                            )
                            if st.session_state.pexels_img_info.get("url"):
                                html_content = insert_pexels_image_html(
                                    html_content,
                                    st.session_state.pexels_img_info,
                                    st.session_state.blog_keyword,
                                )
                            result = post_to_wordpress(post_title, html_content, status="publish")
                            st.success(
                                f"발행 완료! 포스트 ID: {result.get('id')}  \n"
                                f"[글 바로 보기]({result.get('link', '')})"
                            )
                        except requests.HTTPError as e:
                            st.error(f"발행 실패 (HTTP {e.response.status_code}): {e.response.text[:300]}")
                        except Exception as e:
                            st.error(f"발행 실패: {e}")
            with wp_col2:
                if st.button("📝 임시저장(Draft)으로 올리기",
                             use_container_width=True, key="wp_draft"):
                    with st.spinner("임시저장 중..."):
                        try:
                            post_title   = extract_title(st.session_state.blog_result)
                            html_content = md_lib.markdown(
                                st.session_state.blog_result,
                                extensions=["tables", "fenced_code"],
                            )
                            if st.session_state.pexels_img_info.get("url"):
                                html_content = insert_pexels_image_html(
                                    html_content,
                                    st.session_state.pexels_img_info,
                                    st.session_state.blog_keyword,
                                )
                            result = post_to_wordpress(post_title, html_content, status="draft")
                            st.success(
                                f"임시저장 완료! 포스트 ID: {result.get('id')}  \n"
                                f"[글 확인하기]({result.get('link', '')})"
                            )
                        except requests.HTTPError as e:
                            st.error(f"임시저장 실패 (HTTP {e.response.status_code}): {e.response.text[:300]}")
                        except Exception as e:
                            st.error(f"임시저장 실패: {e}")
        else:
            st.info(
                "`.env` 파일에 `WP_URL`, `WP_USERNAME`, `WP_APP_PASSWORD`를 설정하면 "
                "워드프레스 발행 버튼이 활성화됩니다.",
                icon="ℹ️",
            )

    st.stop()


# ══════════════════════════════════════════════════════════════════════════════
# 예약 대시보드 탭
# ══════════════════════════════════════════════════════════════════════════════

if menu == "📅 예약 대시보드":
    st.markdown("""
<div style="background:linear-gradient(120deg,#1e1b4b 0%,#312e81 60%,#4338ca 100%);
            padding:1.2rem 2rem;border-radius:14px;margin-bottom:1.4rem;
            box-shadow:0 4px 24px rgba(67,56,202,.35);">
  <h2 style="margin:0;color:#fff;font-size:1.4rem;font-weight:900;">글 자동 생성 · 예약 관리 대시보드</h2>
  <p style="margin:.25rem 0 0;color:rgba(255,255,255,.80);font-size:.9rem;">
    키워드 큐 관리 → Pexels 이미지 자동 삽입 → 워드프레스 백그라운드 발행
  </p>
</div>
""", unsafe_allow_html=True)

    # ── 키워드 큐 관리 ────────────────────────────────────────────────────────
    st.markdown("#### 키워드 큐 관리")
    add_col1, add_col2 = st.columns([4, 1])
    with add_col1:
        new_kw = st.text_input("키워드 추가", placeholder="SBS아카데미 대전 영상편집 과정",
                               label_visibility="collapsed", key="new_kw_input")
    with add_col2:
        if st.button("+ 큐에 추가", use_container_width=True, key="add_kw"):
            if new_kw.strip():
                q = load_queue()
                q.append({"keyword": new_kw.strip(), "status": "pending",
                           "added_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                           "img_status": "—", "post_id": None,
                           "post_link": "", "error": ""})
                save_queue(q)
                st.rerun()

    with st.expander("📋 여러 키워드 한 번에 추가"):
        bulk_kw = st.text_area("키워드 목록 (한 줄에 하나씩)", height=140,
                               placeholder="SBS아카데미 대전 포토샵 수업\nSBS아카데미 대전 웹디자인 취업\n...",
                               key="bulk_kw")
        if st.button("모두 큐에 추가", key="bulk_add"):
            kws = [k.strip() for k in bulk_kw.strip().splitlines() if k.strip()]
            if kws:
                q = load_queue()
                for kw in kws:
                    q.append({"keyword": kw, "status": "pending",
                               "added_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                               "img_status": "—", "post_id": None,
                               "post_link": "", "error": ""})
                save_queue(q)
                st.rerun()

    # ── 큐 현황 테이블 ────────────────────────────────────────────────────────
    queue = load_queue()
    n_total     = len(queue)
    n_pending   = sum(1 for i in queue if i["status"] == "pending")
    n_done      = sum(1 for i in queue if i["status"] == "done")
    n_fail      = sum(1 for i in queue if i["status"] == "failed")
    n_proc      = sum(1 for i in queue if i["status"] == "processing")

    if queue:
        st.markdown(
            f"**큐 현황** — 전체 {n_total}개 &nbsp;|&nbsp; "
            f"대기 {n_pending} · 처리중 {n_proc} · 완료 {n_done} · 실패 {n_fail}"
        )
        display_cols = ["keyword", "status", "img_status", "post_link", "error", "added_at"]
        col_labels   = ["키워드", "상태", "이미지", "발행 링크", "오류", "추가 시각"]
        df_q = pd.DataFrame(queue)[display_cols]
        df_q.columns = col_labels
        st.dataframe(df_q, use_container_width=True,
                     height=min(38 * (len(df_q) + 2) + 10, 420))

        c1, c2, c3 = st.columns(3)
        with c1:
            if st.button("✅ 완료 항목 삭제", use_container_width=True, key="del_done"):
                save_queue([i for i in queue if i["status"] != "done"])
                st.rerun()
        with c2:
            if st.button("🔁 실패 항목 재시도", use_container_width=True, key="retry_fail"):
                for i in queue:
                    if i["status"] == "failed":
                        i.update({"status": "pending", "error": ""})
                save_queue(queue)
                st.rerun()
        with c3:
            if st.button("🗑️ 전체 초기화", use_container_width=True, key="clear_all_q"):
                save_queue([])
                st.rerun()
    else:
        st.info("큐가 비어 있습니다. 위에서 키워드를 추가해 주세요.")

    # ── 발행 설정 ─────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 발행 설정")
    opt_c1, opt_c2, opt_c3 = st.columns(3)
    with opt_c1:
        interval_min = st.slider(
            "발행 간격", min_value=30, max_value=360, value=60, step=30,
            format="%d분", key="dash_interval",
            help="포스팅 1개 완료 후 다음 포스팅까지 대기 시간",
        )
    with opt_c2:
        dash_include_img = st.checkbox("📷 Pexels 이미지 자동 삽입", value=True, key="dash_img")
    with opt_c3:
        dash_wp_status = st.radio(
            "발행 상태", ["publish", "draft"],
            format_func=lambda x: "즉시 발행" if x == "publish" else "임시저장",
            key="dash_wp_status",
        )

    # ── 제어 버튼 ─────────────────────────────────────────────────────────────
    st.markdown("---")
    btn_c1, btn_c2, btn_c3 = st.columns(3)
    with btn_c1:
        start_disabled = _BG_STATE["running"] or n_pending == 0
        if st.button("🚀 자동 발행 시작", type="primary",
                     use_container_width=True, key="dash_start",
                     disabled=start_disabled):
            if not gemini_api_key:
                st.warning("사이드바에서 Gemini API Key를 입력해 주세요.")
            elif not (WP_URL and WP_USERNAME and WP_APP_PASSWORD):
                st.warning(".env에 WordPress 정보를 설정해 주세요.")
            else:
                t = threading.Thread(
                    target=_bg_worker,
                    args=(interval_min * 60, gemini_api_key,
                          pexels_api_key, dash_wp_status, dash_include_img),
                    daemon=True,
                )
                _BG_STATE["thread"] = t
                t.start()
                st.rerun()
    with btn_c2:
        if st.button("⏹ 중지", use_container_width=True, key="dash_stop",
                     disabled=not _BG_STATE["running"]):
            _BG_STATE["stop"] = True
            st.rerun()
    with btn_c3:
        if st.button("🔄 새로고침", use_container_width=True, key="dash_refresh"):
            st.rerun()

    # ── 실행 상태 표시 & 자동 폴링 ───────────────────────────────────────────
    if _BG_STATE["running"]:
        st.success(
            f"백그라운드 발행 실행 중 — 발행 간격 {interval_min}분 · "
            f"대기 {n_pending}개 남음. 자동으로 화면이 갱신됩니다."
        )
        time.sleep(6)
        st.rerun()

    st.stop()


# ══════════════════════════════════════════════════════════════════════════════
# AI 과정 설계 검색창
# ══════════════════════════════════════════════════════════════════════════════

hint_examples = [
    "마야 과정 최소/최대 설계해줘",
    "웹디자인 취업 루트 짜줘",
    "영상편집 단기 과정 추천",
    "인테리어 풀 커리큘럼",
    "AI 크리에이터 입문",
    "유튜브 수익화 과정",
]

st.markdown("""
<div class="ai-search-wrap">
  <h3>🤖 AI 과정 설계 검색</h3>
  <p>희망 분야와 목표를 자유롭게 입력하면 최소·최대 커리큘럼을 자동으로 설계합니다.</p>
</div>
""", unsafe_allow_html=True)

# 힌트 칩 (클릭하면 입력창에 채워짐)
hint_cols = st.columns(len(hint_examples))
for col, hint in zip(hint_cols, hint_examples):
    with col:
        if st.button(hint, key=f"hint_{hint}", use_container_width=True):
            st.session_state.ai_query = hint

# 검색 입력창
with st.form("ai_search_form", clear_on_submit=False):
    ai_input = st.text_input(
        "검색",
        value=st.session_state.ai_query,
        placeholder="예: 마야 포트폴리오 과정 추천해줘 / 웹디자인 취업 루트",
        label_visibility="collapsed",
    )
    search_btn = st.form_submit_button("🔍 커리큘럼 설계", use_container_width=True, type="primary")

# 데이터 로드 (검색 전에 필요)
courses       = load_courses()
price_map     = course_lookup(courses)
all_courses   = [c["course"] for c in courses]
CATEGORY_MAP    = build_category_map(courses)
COURSE_TO_CAT   = {c: cat for cat, cs in CATEGORY_MAP.items() for c in cs}
# AI 설계 카드 및 합계 표시용: 단가 × 레벨 수 적용된 가격맵
total_price_map = {c: price_map[c] * course_level_count(c) for c in price_map}

# ── 검색 실행 ─────────────────────────────────────────────────────────────────
if search_btn and ai_input.strip():
    st.session_state.ai_query = ai_input.strip()
    result = design(ai_input.strip(), total_price_map)
    st.session_state.ai_result = result
    if result is None:
        st.warning(
            "인식된 분야가 없습니다. "
            "**마야, 웹, 영상, 인테리어, 그래픽, 유튜브, AI, 블렌더, 파이썬** 등의 키워드를 포함해 입력해 주세요."
        )

# ── 설계 결과 카드 표시 ───────────────────────────────────────────────────────
ai_res = st.session_state.ai_result

if ai_res:
    plan_min = ai_res["min"]
    plan_max = ai_res["max"]
    style    = ai_res["style"]

    st.markdown(
        f'<div class="section-title">✨ AI 설계 결과 — {ai_res["field_label"]}</div>',
        unsafe_allow_html=True,
    )

    show_min = style in ("both", "min")
    show_max = style in ("both", "max")
    card_cols = st.columns(2) if (show_min and show_max) else st.columns(1)

    def _course_tags(names: list[str], price_map: dict, dark: bool) -> str:
        tag_cls = "dc-course-tag-dark" if dark else "dc-course-tag"
        return "".join(
            f'<span class="{tag_cls}">{n} <b style="opacity:.7">{price_map.get(n,0):,}원</b></span>'
            for n in names
        )

    def _relation_html(relations: list[tuple], dark: bool) -> str:
        color = "#94a3b8" if dark else "#6b7280"
        rows = ""
        for course, desc in relations:
            # desc 안에 \n이 있으면 줄바꿈 처리
            desc_html = desc.replace("\n", "<br>")
            rows += (
                f'<div class="dc-relation-item" style="color:{color}">'
                f'<b style="color:{"#a5b4fc" if dark else "#e50012"}">{course}</b> — {desc_html}</div>'
            )
        return rows

    # ── 최소 카드 ─────────────────────────────────────────────────────────────
    if show_min:
        col = card_cols[0] if (show_min and show_max) else card_cols[0]
        with col:
            tags  = _course_tags(plan_min["courses"], total_price_map, dark=False)
            rels  = _relation_html(plan_min["relations"], dark=False)
            cnt   = len(plan_min["courses"])
            st.markdown(f"""
<div class="design-card design-card-min">
  <span class="dc-badge-min">⚡ 최소 과정</span>
  <div class="dc-field-label">{ai_res['field_label']}</div>
  <div class="dc-price-min">{fmt_won(plan_min['total'])}</div>
  <div style="color:#6b7280;font-size:.78rem;margin-bottom:.7rem">{cnt}개 과정</div>
  <div style="margin-bottom:.5rem">{tags}</div>
  <div class="dc-intent-min">{plan_min['intent']}</div>
  <div class="dc-relation-title">📌 과목별 설계 의도</div>
  {rels}
</div>
""", unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("⚡ 최소 과정으로 적용하기", key="apply_ai_min",
                         use_container_width=True, type="primary"):
                st.session_state.selected_courses = plan_min["courses"]
                memo_lines = [f"[AI 설계 — {ai_res['field_label']} 최소 과정]", "",
                               plan_min["intent"], "",
                               "■ 과목별 설계 의도"] + \
                             [f"• {c}: {d.split(chr(10))[0]}" for c, d in plan_min["relations"]]
                st.session_state.ai_design_memo = "\n".join(memo_lines)
                st.rerun()

    # ── 최대 카드 ─────────────────────────────────────────────────────────────
    if show_max:
        col = card_cols[1] if (show_min and show_max) else card_cols[0]
        with col:
            tags  = _course_tags(plan_max["courses"], total_price_map, dark=True)
            rels  = _relation_html(plan_max["relations"], dark=True)
            cnt   = len(plan_max["courses"])
            st.markdown(f"""
<div class="design-card design-card-max">
  <span class="dc-badge-max">🚀 최대 과정</span>
  <div class="dc-field-label" style="color:#94a3b8">{ai_res['field_label']}</div>
  <div class="dc-price-max">{fmt_won(plan_max['total'])}</div>
  <div style="color:#64748b;font-size:.78rem;margin-bottom:.7rem">{cnt}개 과정</div>
  <div style="margin-bottom:.5rem">{tags}</div>
  <div class="dc-intent-max">{plan_max['intent']}</div>
  <div class="dc-relation-title" style="color:#94a3b8">📌 과목별 설계 의도</div>
  {rels}
</div>
""", unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🚀 최대 과정으로 적용하기", key="apply_ai_max",
                         use_container_width=True):
                st.session_state.selected_courses = plan_max["courses"]
                memo_lines = [f"[AI 설계 — {ai_res['field_label']} 최대 과정]", "",
                               plan_max["intent"], "",
                               "■ 과목별 설계 의도"] + \
                             [f"• {c}: {d.split(chr(10))[0]}" for c, d in plan_max["relations"]]
                st.session_state.ai_design_memo = "\n".join(memo_lines)
                st.rerun()

    # 초기화 버튼
    if st.button("✕ 설계 결과 닫기", key="clear_ai_result"):
        st.session_state.ai_result = None
        st.session_state.ai_design_memo = ""
        st.rerun()

    st.markdown("---")


# ── 프리셋 트리거 처리 ────────────────────────────────────────────────────────
if st.session_state.preset_trigger:
    preset_courses = PRESETS[st.session_state.preset_trigger]["courses"]
    st.session_state.selected_courses = [c for c in preset_courses if c in price_map]
    st.session_state.preset_trigger = None


# ── 레이아웃: 과정 선택 | 합계 카드 ──────────────────────────────────────────
col_left, col_right = st.columns([3, 1.4], gap="large")

with col_left:
    st.markdown('<div class="section-title">📚 학과 선택</div>', unsafe_allow_html=True)

    all_cats = list(CATEGORY_MAP.keys())
    sel_depts = st.multiselect(
        "학과를 선택하세요 (복수 선택 가능)",
        options=all_cats,
        default=all_cats,
        label_visibility="collapsed",
    )

    # 선택 학과에 속하는 과정 필터
    filtered_courses = [
        c for c in all_courses
        if COURSE_TO_CAT.get(c, "기타") in (sel_depts or all_cats)
    ]

    # 현재 세션의 selected_courses 중 filtered에 있는 것만 default로
    current_valid = [c for c in st.session_state.selected_courses if c in filtered_courses]

    st.markdown('<div class="section-title">✅ 상세 과정 선택</div>', unsafe_allow_html=True)
    chosen = st.multiselect(
        "수강할 과정을 선택하세요",
        options=filtered_courses,
        default=current_valid,
        label_visibility="collapsed",
        key="course_multiselect",
    )
    st.session_state.selected_courses = chosen


with col_right:
    # s_discount는 사이드바 슬라이더에서 직접 가져옴
    n        = sum(course_level_count(c) for c in st.session_state.selected_courses)
    n_titles = len(st.session_state.selected_courses)

    df         = build_df(st.session_state.selected_courses, price_map, s_discount)
    subtotal   = int(df["_price"].sum())  if not df.empty else 0
    after_pct  = int(df["_final"].sum())  if not df.empty else 0
    fixed_disc = int(review_won) + int(ddaza_won)
    total      = max(0, after_pct - fixed_disc)
    savings    = subtotal - total

    st.markdown('<div class="section-title">💰 합계</div>', unsafe_allow_html=True)

    if ddaza_on and after_pct < 1_500_000:
        st.warning("따즈아 할인은 할인 적용 후 150만원 이상 시 가능합니다.")

    st.markdown(f"""
    <div class="total-card">
      <div class="total-label">정가 합계</div>
      <div class="total-amount">{fmt_won(subtotal)}</div>
      <div class="total-savings">
        복합 할인율 {s_discount:.1f}% 적용 → 최종 {fmt_won(total)}<br>
        절약 {fmt_won(savings)}
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.metric("선택 과정 수", f"{n_titles}종 ({n}단계)")
    st.metric("정가 합계",   fmt_won(subtotal))

    # ── 단계별 할인 미리보기 ──────────────────────────────────────────────────
    _disc_short = ["8개↑할인", "따따아등록", "첫결제", "원장권한", "기타"]
    if subtotal > 0 and any([disc_rate1, disc_rate2, disc_rate3, disc_rate4, disc_rate5]):
        _p = subtotal
        _rows = [f"**정가** → {fmt_won(_p)}"]
        for _dr, _lbl in zip([disc_rate1, disc_rate2, disc_rate3, disc_rate4, disc_rate5], _disc_short):
            if _dr:
                _p = round(_p * (1 - _dr / 100))
                _rows.append(f"_{_lbl}_ -{_dr}% → **{fmt_won(_p)}**")
        with st.expander("📊 단계별 할인 미리보기", expanded=True):
            for _r in _rows:
                st.markdown(_r)

    if fixed_disc:
        st.metric("% 할인 후", fmt_won(after_pct))
        st.metric("추가 금액 할인", f"-{fmt_won(fixed_disc)}", delta_color="inverse")
    st.metric("최종 금액",   fmt_won(total),   delta=f"-{fmt_won(savings)}" if savings else None,
              delta_color="inverse")


# ── 수강료 테이블 ─────────────────────────────────────────────────────────────
st.markdown('<div class="section-title">📋 선택 과정 수강료 내역</div>', unsafe_allow_html=True)

if df.empty:
    st.info("위에서 과정을 선택하면 수강료 내역이 자동으로 표시됩니다.")
else:
    display_df = df[["과정명","단가","단계","정가","할인가","절약"]].reset_index(drop=True)
    display_df.index += 1
    st.dataframe(
        display_df,
        use_container_width=True,
        height=min(38 * (len(display_df) + 2) + 10, 500),
    )


# ── 최소/최대 추천 커리큘럼 프리셋 ────────────────────────────────────────────
st.markdown('<div class="section-title">🎯 추천 커리큘럼 프리셋</div>', unsafe_allow_html=True)

preset_cols = st.columns(len(PRESETS))
for col, (label, info) in zip(preset_cols, PRESETS.items()):
    with col:
        st.markdown(f"<small style='color:#6b7280'>{info['desc']}</small>", unsafe_allow_html=True)
        if st.button(label, use_container_width=True, key=f"preset_{label}"):
            st.session_state.preset_trigger = label
            st.rerun()


# ── 저장 섹션 ─────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown('<div class="section-title">💾 상담일지 저장</div>', unsafe_allow_html=True)

if df.empty:
    st.warning("과정을 선택한 뒤 저장할 수 있습니다.")
else:
    today_str   = datetime.now().strftime("%Y-%m-%d %H:%M")
    date_file   = datetime.now().strftime("%Y%m%d_%H%M")
    client_name = s_name.strip() or "미입력"

    md_content   = build_md(client_name, s_contact, s_field, s_consultant, s_memo,
                             df, total, subtotal, savings, s_discount, today_str)
    html_content = build_html(client_name, s_contact, s_field, s_consultant, s_memo,
                               df, total, subtotal, savings, float(s_discount), today_str)
    excel_bytes  = build_excel(
        client_name, s_contact, s_field, s_consultant, s_memo,
        df, total, subtotal, savings, float(s_discount), today_str,
        float(disc_rate1), disc_reason1,
        float(disc_rate2), disc_reason2,
        float(disc_rate3), disc_reason3,
        float(disc_rate4), disc_reason4,
        float(disc_rate5), disc_reason5,
    )

    btn_col1, btn_col2, btn_col3, btn_col4 = st.columns(4)

    # ① MD 저장 (파일 + 다운로드)
    with btn_col1:
        if st.button("📄 상담일지 MD 저장", use_container_width=True, type="primary"):
            fname = LOGS_DIR / f"상담일지_{client_name}_{date_file}.md"
            fname.write_text(md_content, encoding="utf-8")
            st.markdown(f'<div class="ok-banner">✅ 저장 완료: <code>{fname.name}</code></div>',
                        unsafe_allow_html=True)

        st.download_button(
            "⬇️ MD 다운로드",
            data=md_content.encode("utf-8"),
            file_name=f"상담일지_{client_name}_{date_file}.md",
            mime="text/markdown",
            use_container_width=True,
        )

    # ② HTML 저장 (브라우저에서 인쇄 → PDF 가능)
    with btn_col2:
        if st.button("🌐 HTML 보고서 저장", use_container_width=True, type="primary"):
            fname = LOGS_DIR / f"상담일지_{client_name}_{date_file}.html"
            fname.write_text(html_content, encoding="utf-8")
            st.markdown(f'<div class="ok-banner">✅ 저장 완료: <code>{fname.name}</code><br>'
                        f'<small>브라우저에서 열고 Ctrl+P → PDF로 저장하세요.</small></div>',
                        unsafe_allow_html=True)

        st.download_button(
            "⬇️ HTML 다운로드",
            data=html_content.encode("utf-8"),
            file_name=f"상담일지_{client_name}_{date_file}.html",
            mime="text/html",
            use_container_width=True,
        )

    # ③ 두 파일 동시 저장
    with btn_col3:
        if st.button("💾 MD + HTML 동시 저장", use_container_width=True):
            for ext, content, enc in [
                ("md",   md_content,   "utf-8"),
                ("html", html_content, "utf-8"),
            ]:
                p = LOGS_DIR / f"상담일지_{client_name}_{date_file}.{ext}"
                p.write_text(content, encoding=enc)
            st.markdown(
                f'<div class="ok-banner">✅ 두 파일 저장 완료<br>'
                f'<code>logs/상담일지_{client_name}_{date_file}.md</code><br>'
                f'<code>logs/상담일지_{client_name}_{date_file}.html</code></div>',
                unsafe_allow_html=True,
            )
        st.caption("logs/ 폴더에 두 형식 모두 저장됩니다.")

    # ④ 엑셀 견적서 다운로드
    with btn_col4:
        st.download_button(
            "📊 엑셀 견적서 다운로드",
            data=excel_bytes,
            file_name=f"견적서_{client_name}_{date_file}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )
        st.caption("SBS컴퓨터아트학원 양식으로 저장됩니다.")


# ══════════════════════════════════════════════════════════════════════════════
# AI 멘토링 어시스턴트
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("---")
st.markdown("""
<style>
.mentor-wrap {
    background:linear-gradient(135deg,#0f172a 0%,#1a1a2e 100%);
    border-radius:16px; padding:1.5rem 1.8rem; margin-bottom:1.4rem;
    border:1px solid #374151;
    box-shadow:0 4px 24px rgba(0,0,0,.25);
}
.mentor-wrap h3 { color:#fbbf24; font-size:1rem; font-weight:800; margin:0 0 .15rem; }
.mentor-wrap p  { color:#94a3b8; font-size:.82rem; margin:0 0 .8rem; }
.mentor-answer {
    background:#1e293b; border:1.5px solid #334155;
    border-radius:12px; padding:1.2rem 1.4rem;
    color:#e2e8f0; font-size:.9rem; line-height:1.7;
    margin-top:.8rem; white-space:pre-wrap;
}
</style>
<div class="mentor-wrap">
  <h3>🧑‍🏫 AI 멘토링 어시스턴트</h3>
  <p>SBS아카데미 10년 차 베테랑 커리어 멘토에게 무엇이든 물어보세요.</p>
</div>
""", unsafe_allow_html=True)

# 추천 질문 버튼
suggest_cols = st.columns(2)
for col, q in zip(suggest_cols, ["비전공자 취업 준비", "에셋과 에펙의 차이"]):
    with col:
        if st.button(q, key=f"suggest_{q}", use_container_width=True):
            st.session_state.mentor_query = q

with st.form("mentor_form", clear_on_submit=False):
    mentor_input = st.text_input(
        "질문",
        value=st.session_state.mentor_query,
        placeholder="예: 비전공자도 그래픽 디자인 취업이 가능한가요?",
        label_visibility="collapsed",
    )
    mentor_btn = st.form_submit_button("💬 멘토에게 질문하기", use_container_width=True, type="primary")

if mentor_btn and mentor_input.strip():
    st.session_state.mentor_query = mentor_input.strip()
    if not gemini_api_key:
        st.warning("사이드바에서 Gemini API Key를 입력해 주세요.")
    else:
        with st.spinner("멘토가 생각 중입니다..."):
            try:
                genai.configure(api_key=gemini_api_key)
                # v1beta 명시 없이 기본 안정 버전(v1) 사용
                model = genai.GenerativeModel(
                    model_name="gemini-2.5-flash",
                    system_instruction="너는 SBS컴퓨터아트학원의 10년 차 베테랑 커리어 멘토야.",
                )
                response = model.generate_content(mentor_input.strip())
                st.session_state.mentor_answer = response.text
            except Exception as e:
                err_str = str(e)
                if "429" in err_str or "quota" in err_str.lower() or "resource" in err_str.lower():
                    st.session_state.mentor_answer = "⏳ QUOTA_EXCEEDED"
                else:
                    st.session_state.mentor_answer = f"❌ 오류가 발생했습니다: {e}"

if st.session_state.mentor_answer:
    if st.session_state.mentor_answer == "⏳ QUOTA_EXCEEDED":
        st.markdown("""
<div style="background:#fefce8;border:1.5px solid #ca8a04;border-radius:12px;
            padding:1.2rem 1.4rem;color:#92400e;font-size:.95rem;line-height:1.7;">
  ⏳ <b>현재 사용자가 많아 잠시 후 다시 시도해 주세요.</b><br>
  <span style="font-size:.85rem;opacity:.8">AI 멘토링 서비스 사용량이 일시적으로 초과되었습니다. 약 1분 후 재시도해 주세요.</span>
</div>
""", unsafe_allow_html=True)
    else:
        st.markdown("**🧑‍🏫 멘토의 답변**")
        st.markdown(
            f'<div class="mentor-answer">{st.session_state.mentor_answer}</div>',
            unsafe_allow_html=True,
        )
    if st.button("✕ 닫기", key="clear_mentor"):
        st.session_state.mentor_answer = ""
        st.rerun()


# ── 푸터 ──────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    "<p style='text-align:center;color:#9ca3af;font-size:.8rem'>"
    "SBS아카데미 대전지점 상담 자동화 시스템 &nbsp;|&nbsp; "
    "데이터 출처: 수강료 엑셀 (2026.04.01 기준)</p>",
    unsafe_allow_html=True,
)
