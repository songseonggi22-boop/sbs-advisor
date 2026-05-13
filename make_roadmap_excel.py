import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.cell import MergedCell
from copy import copy

CATEGORIES = {
    'AI 에이전트': [
        ('AI 프롬프트엔지니어링 1', 600000),
        ('AI 프롬프트엔지니어링 2', 600000),
        ('AI 에이전트 기초',         600000),
        ('AI 에이전트 심화',         600000),
    ],
    '인공지능융합 SW': [
        ('파이썬 1',400000),('파이썬 2',400000),('파이썬 3',400000),
        ('자바 1',400000),('자바 2',400000),
        ('C언어 1',400000),('C언어 2',400000),
        ('머신러닝 1',400000),('머신러닝 2',400000),('머신러닝 3',400000),('머신러닝 4',400000),
    ],
    'UIUX·웹': [
        ('발상과표현',400000),('색채학',400000),('포토샵, 일러스트',350000),('포토웍스, 디일러',400000),
        ('인포그래픽',400000),('웹 1 (HTML5/CSS3)',400000),('웹 2 (J-query/JS)',400000),
        ('웹 3 (반응형)',400000),('UIUX 1 (피그마 기초)',400000),('UIUX 2 (UX분석)',400000),
        ('UIUX 3 (UI제작)',400000),('AI 바이브코딩-웹',500000),
        ('웹 포트폴리오 1',500000),('웹 포트폴리오 2',500000),
    ],
    '시각편집디자인': [
        ('발상과표현',400000),('색채학',400000),('포토샵, 일러스트',350000),('포토웍스, 디일러',400000),
        ('인포그래픽',400000),('인디자인',350000),('그래픽아트웍',400000),
        ('AI 크리에이터-시각편집',500000),('편집 포트폴리오 1',500000),('편집 포트폴리오 2',500000),
    ],
    '인테리어디자인': [
        ('발상과표현',400000),('색채학',400000),('포토샵, 일러스트',350000),('포토웍스, 디일러',400000),
        ('캐드 1',400000),('캐드 2 (도면제작/CAT)',400000),
        ('스케치업 1 (Modeling)',400000),('스케치업 2 (Rendering)',400000),
        ('맥스 1 (Modeling)',450000),('맥스 2 (Modeling+)',450000),('맥스 3 (Rendering)',450000),
        ('인테리어 스케치 1',400000),('인테리어 스케치 2',400000),
        ('인테리어 스케치 3',400000),('인테리어 스케치 4',400000),
        ('시공실무',300000),('실내건축이론',400000),
        ('실내건축자격증 1',400000),('실내건축자격증 2',400000),
        ('실내건축자격증 3',400000),('실내건축자격증 4',400000),
        ('AI 크리에이터-인테리어',500000),
        ('인테리어 포트폴리오 1',500000),('인테리어 포트폴리오 2',500000),('인테리어 포트폴리오 3',500000),
    ],
    '모션그래픽': [
        ('발상과표현',400000),('색채학',400000),('포토샵, 일러스트',350000),('포토웍스, 디일러',400000),
        ('인포그래픽',400000),('그래픽아트웍',400000),
        ('프리미어',400000),('에펙',450000),('모션 에펙',450000),('어드벤스 에펙',450000),
        ('블렌더 1',500000),('블렌더 2',500000),
        ('시네마4D 1',500000),('시네마4D 2',500000),('시네마4D 3',500000),('시네마4D 4',500000),
        ('AI 크리에이터-영상모션',500000),
        ('모션 포트폴리오 1',600000),('모션 포트폴리오 2',600000),('모션 포트폴리오 3',600000),
    ],
    'CG·VFX (마야)': [
        ('발상과표현',400000),('색채학',400000),('해부학',400000),
        ('포토샵, 일러스트',350000),('포토웍스, 디일러',400000),
        ('프리미어',400000),('에펙',450000),('모션 에펙',450000),('어드벤스 에펙',450000),
        ('블렌더 1',500000),('블렌더 2',500000),
        ('마야 1 (기초&모델링)',600000),('마야 2 (렌더링/텍스쳐)',600000),
        ('마야 3 (리깅/언리얼)',600000),('마야 4 (컴포지팅/트랙킹)',600000),
        ('마야 5 (모델링심화/Z-Brush)',600000),('마야 6 (텍스쳐링/합성)',600000),
        ('마야 7 (리깅/애니메이션)',600000),
        ('AI 크리에이터-CG마야',500000),
        ('마야 포트폴리오 VFX 1',600000),('마야 포트폴리오 VFX 2',600000),
    ],
    '웹툰': [
        ('발상과표현',400000),('색채학',400000),('포토샵, 일러스트',350000),('포토웍스, 디일러',400000),
        ('웹툰 1 (캐릭터 IP개발)',500000),('웹툰 2 (색채·캐릭터)',500000),
        ('웹툰 3 (연출·콘티)',500000),('웹툰 4 (배경 제작)',500000),
        ('웹툰 5 (채색심화)',500000),('웹툰 6 (스토리보드)',500000),
        ('AI 크리에이터-아트웍',500000),
    ],
    '디지털드로잉 (아트웍)': [
        ('발상과표현',400000),('색채학',400000),('해부학',400000),
        ('포토샵, 일러스트',350000),('포토웍스, 디일러',400000),
        ('디지털드로잉 1 (타블렛 기초)',500000),('디지털드로잉 2 (구도·얼굴묘사)',500000),
        ('디지털드로잉 3 (채색 기초)',500000),('디지털드로잉 4 (채색심화)',500000),
        ('디지털드로잉 5 (반창작)',500000),('디지털드로잉 6 (투시도·창작)',500000),
        ('AI 크리에이터-아트웍',500000),
        ('드로잉 포트폴리오 1',600000),('드로잉 포트폴리오 2',600000),
        ('드로잉 포트폴리오 3',600000),('드로잉 포트폴리오 4',600000),
    ],
}

wb_src = openpyxl.load_workbook('견적서_양식_backup.xlsx.xlsx')
ws_src = wb_src.worksheets[0]
src_heights = {r: d.height for r, d in ws_src.row_dimensions.items()}
DATA_H = src_heights.get(8, 19.5)

def copy_cell(src, dst):
    if isinstance(src, MergedCell):
        return
    dst.value = src.value
    if src.has_style:
        dst.font      = copy(src.font)
        dst.fill      = copy(src.fill)
        dst.border    = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format

def set_val(ws, col_idx, row, val):
    c = ws.cell(row=row, column=col_idx)
    if not isinstance(c, MergedCell):
        c.value = val

def try_merge(ws, range_str):
    try:
        ws.merge_cells(range_str)
    except Exception:
        pass

wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)

for cat_name, courses in CATEGORIES.items():
    n = len(courses)
    DS = 8                   # data start
    DE = DS + n - 1          # data end
    SR = DE + 1              # sum row
    D1 = SR + 3              # discount 1 (SR+1=할인헤더, SR+2=열헤더, SR+3=할인1)
    D2 = D1 + 1
    D3 = D1 + 2
    TD = D3 + 1              # total discount
    NF = TD + 1              # net fee
    TR = NF + 1              # total reg
    I3 = TR + 1              # 3-month
    I6 = I3 + 1              # 6-month
    PI = I6 + 1              # per-item

    ws = wb_out.create_sheet(title=cat_name)

    for col_l, dim in ws_src.column_dimensions.items():
        ws.column_dimensions[col_l].width = dim.width

    # 헤더 행 1~7 복사
    for sr in range(1, 8):
        ws.row_dimensions[sr].height = src_heights.get(sr, DATA_H)
        for c in range(1, ws_src.max_column + 1):
            copy_cell(ws_src.cell(row=sr, column=c), ws.cell(row=sr, column=c))

    # 데이터 행
    for i, (name, price) in enumerate(courses):
        r = DS + i
        ws.row_dimensions[r].height = DATA_H
        for c in range(1, ws_src.max_column + 1):
            copy_cell(ws_src.cell(row=8, column=c), ws.cell(row=r, column=c))
        set_val(ws, 2, r, name)
        set_val(ws, 6, r, price)
        set_val(ws, 4, r, 1)

    # 합계 행
    ws.row_dimensions[SR].height = src_heights.get(10, 33.6)
    for c in range(1, ws_src.max_column + 1):
        copy_cell(ws_src.cell(row=10, column=c), ws.cell(row=SR, column=c))
    ws.cell(row=SR, column=4).value = f'=SUM(D{DS}:D{DE})'
    ws.cell(row=SR, column=6).value = f'=SUM(F{DS}:F{DE})'

    # 할인 헤더 행 (SR+1, SR+2)
    for offset, src_r in [(1, 11), (2, 12)]:
        dr = SR + offset
        ws.row_dimensions[dr].height = src_heights.get(src_r, 19.5)
        for c in range(1, ws_src.max_column + 1):
            copy_cell(ws_src.cell(row=src_r, column=c), ws.cell(row=dr, column=c))

    # 할인 3행
    for idx, src_r in enumerate([13, 14, 15]):
        dr = D1 + idx
        ws.row_dimensions[dr].height = src_heights.get(src_r, 19.5)
        for c in range(1, ws_src.max_column + 1):
            copy_cell(ws_src.cell(row=src_r, column=c), ws.cell(row=dr, column=c))
    ws.cell(row=D1, column=5).value = f'=F{SR}*D{D1}'
    ws.cell(row=D2, column=5).value = f'=(F{SR}*(1-D{D1})*D{D2})'
    ws.cell(row=D3, column=5).value = f'=(F{SR}*(1-D{D1})*(1-D{D2})*D{D3})'

    # 합계 섹션
    for idx, src_r in enumerate([16, 17, 18, 19, 20, 21]):
        dr = TD + idx
        ws.row_dimensions[dr].height = src_heights.get(src_r, 23.25)
        for c in range(1, ws_src.max_column + 1):
            copy_cell(ws_src.cell(row=src_r, column=c), ws.cell(row=dr, column=c))
    ws.cell(row=TD, column=4).value = f'=ROUNDUP(SUM(E{D1}:G{D3}),-3)'
    ws.cell(row=NF, column=4).value = f'=F{SR}-D{TD}'
    ws.cell(row=TR, column=4).value = f'=SUM(D{NF}+C{SR})'
    ws.cell(row=I3, column=4).value = f'=SUM(D{TR}/3)'
    ws.cell(row=I6, column=4).value = f'=SUM(D{TR}/6)'
    ws.cell(row=PI, column=4).value = f'=SUM(D{NF}/D{SR})'

    # 안내사항
    for idx, src_r in enumerate(range(22, 35)):
        dr = PI + 1 + idx
        ws.row_dimensions[dr].height = src_heights.get(src_r, 16.5)
        for c in range(1, ws_src.max_column + 1):
            copy_cell(ws_src.cell(row=src_r, column=c), ws.cell(row=dr, column=c))
    ws.cell(row=PI + 1 + (31 - 22), column=1).value = None  # 계좌 삭제

    # SBS 정보
    ws['A1'] = f'{cat_name} 교육과정 수강료 안내'
    ws['F5'] = '안재현 교육멘토'
    ws['F6'] = '042-710-8921'

    # 병합셀
    for m in ['A1:G3','F4:G4','B7:C7','F7:G7']:
        try_merge(ws, m)
    for r in range(DS, DE + 1):
        try_merge(ws, f'B{r}:C{r}')
        try_merge(ws, f'F{r}:G{r}')
    try_merge(ws, f'F{SR}:G{SR}')
    try_merge(ws, f'F{SR+1}:G{SR+1}')
    try_merge(ws, f'A{SR+2}:C{SR+2}')
    try_merge(ws, f'E{SR+2}:G{SR+2}')
    try_merge(ws, f'A{D1}:A{D3}')
    for r in [D1, D2, D3]:
        try_merge(ws, f'B{r}:C{r}')
        try_merge(ws, f'E{r}:G{r}')
    for r in [TD, NF, TR, I3, I6, PI]:
        try_merge(ws, f'A{r}:C{r}')
        try_merge(ws, f'D{r}:G{r}')

    print(f'  [{cat_name}] {n}과목 완료')

wb_out.calculation.calcMode = 'auto'
wb_out.calculation.fullCalcOnLoad = True
wb_out.save('SBS_로드맵_견적서.xlsx')
print('\nSBS_로드맵_견적서.xlsx 저장 완료!')
