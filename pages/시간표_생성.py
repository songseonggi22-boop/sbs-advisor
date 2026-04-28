"""
pages/시간표_생성.py — SBS아카데미 과정 검색 & 개인 시간표 생성
데이터 소스: SBS평일/*.xls, SBS주말/*.xls (HTML 형식 강의시간표)
"""

from __future__ import annotations
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
import io, re

import pandas as pd
import streamlit as st
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

ROOT           = Path(__file__).parent.parent
# sbs시간표 최신화/ 를 단일 데이터 소스로 사용
# 폴더가 없으면 레거시 SBS평일/ 폴더로 폴백
_LATEST_DIR    = ROOT / "sbs시간표 최신화"
WEEKDAY_DIR    = _LATEST_DIR if _LATEST_DIR.exists() else ROOT / "SBS평일"
WEEKEND_DIR    = ROOT / "SBS주말"   # optional — 없어도 에러 없음
TEMPLATE_PATH  = ROOT / "★전체시간표 (양식).xlsx"

# ══════════════════════════════════════════════════════════════════════════════
# 데이터 모델
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class CourseEntry:
    name:       str          # 과정명
    room:       str          # 강의실 (A-1 등)
    start_time: str          # 시작 시간 (09:00)
    end_time:   str          # 종료 시간 (11:00)
    instructor: str          # 강사명
    days:       str          # 수업 요일 (월~목, 토/일 등)
    start_date: str          # 개강일 YYYY-MM-DD
    end_date:   str          # 종강일 YYYY-MM-DD
    capacity:   str          # 정원
    enrolled:   str          # 배정
    sheet:      str          # "평일" / "주말"
    source:     str          # 파일명

    @property
    def period_label(self) -> str:
        """2026-04-13 ~ 2026-05-11 형태."""
        if self.start_date and self.end_date:
            return f"{self.start_date}  ~  {self.end_date}"
        return self.start_date or self.end_date or "-"

    @property
    def month_label(self) -> str:
        """4월 ~ 5월 형태."""
        try:
            s = datetime.strptime(self.start_date, "%Y-%m-%d")
            e = datetime.strptime(self.end_date,   "%Y-%m-%d")
            if s.month == e.month:
                return f"{s.year}년 {s.month}월"
            return f"{s.year}년 {s.month}월 ~ {e.month}월"
        except Exception:
            return self.period_label


# ══════════════════════════════════════════════════════════════════════════════
# 파싱
# ══════════════════════════════════════════════════════════════════════════════

_DAY_ORDER = ['월', '화', '수', '목', '금', '토', '일']

def _days_to_set(days: str) -> set[str]:
    """요일 문자열 → 개별 요일 집합.
    월~목 → {월,화,수,목} / 화/목 → {화,목} / 월수금 → {월,수,금}
    """
    if not days:
        return set()
    result: set[str] = set()
    if '~' in days:
        parts = days.split('~')
        if len(parts) == 2:
            try:
                s = _DAY_ORDER.index(parts[0])
                e = _DAY_ORDER.index(parts[1])
                for i in range(min(s, e), max(s, e) + 1):
                    result.add(_DAY_ORDER[i])
            except ValueError:
                pass
    else:
        for c in days.replace('/', ''):
            if c in _DAY_ORDER:
                result.add(c)
    return result


def _days_badge_html(days: str) -> str:
    """월화수목금토일 각각을 활성/비활성 뱃지로 렌더링."""
    active = _days_to_set(days)
    parts = []
    for d in _DAY_ORDER:
        if d in active:
            parts.append(
                f"<span style='display:inline-flex;align-items:center;justify-content:center;"
                f"width:1.45rem;height:1.45rem;border-radius:50%;"
                f"background:#1F4E79;color:#fff;font-size:.72rem;font-weight:800;"
                f"margin:0 .08rem;'>{d}</span>"
            )
        else:
            parts.append(
                f"<span style='display:inline-flex;align-items:center;justify-content:center;"
                f"width:1.45rem;height:1.45rem;border-radius:50%;"
                f"background:#e2e8f0;color:#94a3b8;font-size:.72rem;font-weight:500;"
                f"margin:0 .08rem;'>{d}</span>"
            )
    return ''.join(parts)


def _calc_end_time(start_time: str, rowspan: int) -> str:
    """시작 시간(HH:MM) + rowspan × 30분 → 종료 시간."""
    from datetime import timedelta
    try:
        dt = datetime.strptime(start_time, "%H:%M")
        dt += timedelta(minutes=30 * rowspan)
        return dt.strftime("%H:%M")
    except Exception:
        return ''


def _parse_cell_text(text: str) -> tuple[str, str, str, str, str] | None:
    """
    셀 텍스트 → (과정명, 강사, 요일, 개강일, 종강일) or None
    예) "에펙전체출석율 : 60%...배정:28윤진3월~목개:2026-04-13종:2026-05-11"
    """
    text = text.strip()
    if not text:
        return None

    # 과정명: "전체출석율" 또는 "수업없음" 앞
    m_name = re.match(r'^(.+?)(?:전체출석율|수업없음)', text)
    if not m_name:
        return None
    name = m_name.group(1).strip()
    # /주말 제거
    name = re.sub(r'/주말.*', '', name).strip()
    if not name or len(name) > 40:
        return None
    # 숫자·특수문자만인 것 제외
    if re.fullmatch(r'[\d\W]+', name):
        return None

    # 개강·종강일
    m_start = re.search(r'개:(\d{4}-\d{2}-\d{2})', text)
    m_end   = re.search(r'종:(\d{4}-\d{2}-\d{2})', text)
    start_date = m_start.group(1) if m_start else ''
    end_date   = m_end.group(1)   if m_end   else ''

    # 요일: 개: 바로 앞에 있는 요일 문자열 추출
    # 예) 월~목개: / 화/목개: / 월수금월수개: / 금개:
    m_days = re.search(r'([월화수목금토일][월화수목금토일~\/]*)개:', text)
    raw_days = m_days.group(1) if m_days else ''

    # 중복 요일 제거: 월수금월수 → 월수금
    if raw_days and '~' not in raw_days and '/' not in raw_days:
        seen_d: set[str] = set()
        deduped: list[str] = []
        for c in raw_days:
            if c not in seen_d:
                seen_d.add(c)
                deduped.append(c)
        days = ''.join(deduped)
    else:
        days = raw_days

    # 강사명: 요일 문자열 직전의 한글 2~5자 (선택적)
    instructor = ''
    if raw_days:
        days_pos = text.rfind(raw_days + '개:')
        if days_pos > 0:
            before = text[:days_pos]
            m_instr = re.search(r'([가-힣]{2,5})\s*\d*\s*$', before)
            if m_instr and m_instr.group(1) not in ('재직자', '수업없음', '정원', '배정'):
                instructor = m_instr.group(1)

    return name, instructor, days, start_date, end_date


def _load_xls_file(fp: Path, sheet: str) -> list[CourseEntry]:
    """HTML 형식 .xls 파일 하나 파싱 → CourseEntry 리스트."""
    entries: list[CourseEntry] = []
    try:
        raw  = fp.read_bytes().decode('utf-8', errors='replace')
        soup = BeautifulSoup(raw, 'html.parser')
        tbl  = soup.find('table')
        if not tbl:
            return entries

        rows = tbl.find_all('tr')
        if len(rows) < 4:
            return entries

        # 헤더 행들에서 강의실(서브룸) 이름 추출 (행2 기준)
        room_map: dict[int, str] = {}   # col_index → room_name
        sub_row = rows[1] if len(rows) > 1 else rows[0]
        col_idx = 0
        for td in sub_row.find_all(['td', 'th']):
            span = int(td.get('colspan', 1))
            val  = td.get_text(strip=True)
            if val and val not in ('', '정원'):
                for k in range(span):
                    room_map[col_idx + k] = val
            col_idx += span

        # 데이터 행 (행4 이후: 시간 + 강의실별 과정)
        cur_time = ''
        for row in rows[3:]:
            cells = row.find_all(['td', 'th'])
            if not cells:
                continue

            # 첫 번째 셀: 시간
            first_val = cells[0].get_text(strip=True)
            if re.match(r'^\d{2}:\d{2}$', first_val):
                cur_time = first_val

            # 나머지 셀들
            data_col = 1   # room_map 인덱스: row1의 첫 셀(빈 코너셀)이 col0이므로 1부터 시작
            for i, td in enumerate(cells[1:], start=1):
                span    = int(td.get('colspan', 1))
                rowspan = int(td.get('rowspan', 1))
                text    = td.get_text(strip=True)
                room    = room_map.get(data_col, f'Col{data_col}')
                parsed  = _parse_cell_text(text)
                if parsed:
                    name, instr, days, sd, ed = parsed
                    end_time = _calc_end_time(cur_time, rowspan)
                    entries.append(CourseEntry(
                        name=name, room=room,
                        start_time=cur_time,
                        end_time=end_time,
                        instructor=instr, days=days,
                        start_date=sd, end_date=ed,
                        capacity='', enrolled='',
                        sheet=sheet, source=fp.name,
                    ))
                data_col += span

    except Exception:
        pass
    return entries


@st.cache_resource(show_spinner=False)
def load_all_courses() -> list[CourseEntry]:
    """WEEKDAY_DIR(sbs시간표 최신화 or SBS평일), SBS주말 폴더의 모든 .xls 파일 파싱."""
    all_entries: list[CourseEntry] = []

    for folder, sheet in [(WEEKDAY_DIR, '평일'), (WEEKEND_DIR, '주말')]:
        if not folder.exists():
            continue
        for fp in sorted(folder.glob('*.xls')):
            all_entries.extend(_load_xls_file(fp, sheet))

    # 중복 제거: (과정명+강의실+시간+개강일) 기준
    seen: set[tuple] = set()
    unique: list[CourseEntry] = []
    for e in all_entries:
        key = (e.name, e.room, e.start_time, e.start_date)
        if key not in seen:
            seen.add(key)
            unique.append(e)

    # 개강일 기준 정렬
    unique.sort(key=lambda x: (x.start_date, x.name))
    return unique


def search_courses(entries: list[CourseEntry], keyword: str) -> list[CourseEntry]:
    """과정명 포함 검색 (대소문자·공백 무시)."""
    kw = keyword.strip().lower().replace(' ', '')
    if not kw:
        return []
    return [e for e in entries if kw in e.name.lower().replace(' ', '')]


# ══════════════════════════════════════════════════════════════════════════════
# 수강료 로드 & 매칭
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def load_price_map() -> dict:
    """수강료.xlsx → {과정명: {'평일': 가격, '주말': 가격}} 딕셔너리."""
    fp = ROOT / "수강료.xlsx"
    if not fp.exists():
        return {}

    pm: dict[str, dict[str, int]] = {}

    def _register(raw_name: str, price_val, sheet: str) -> None:
        if not raw_name or not isinstance(price_val, (int, float)) or price_val <= 0:
            return
        # 방학 특강은 별도 키 '방학'으로 등록 → 일반 수강료 덮어쓰기 방지
        is_vacation = '방학' in raw_name
        base = re.sub(r'/주말|/방학|\(방학\)', '', raw_name).strip()
        effective_sheet = '방학' if is_vacation else sheet
        # 쉼표 구분 과목 개별 등록 ("포토샵, 일러스트" → 포토샵, 일러스트 각각)
        for part in [p.strip() for p in base.split(',')]:
            if part:
                # setdefault: 먼저 등록된 가격 우선 (방학 가격으로 덮어쓰기 방지)
                pm.setdefault(part, {}).setdefault(effective_sheet, int(price_val))

    try:
        wb  = openpyxl.load_workbook(fp, data_only=True)
        ws  = wb['수강료'] if '수강료' in wb.sheetnames else wb.active
        cur = '평일'   # 현재 구분 추적

        for row in ws.iter_rows(values_only=True):
            # ── 왼쪽 테이블 (B=1, C=2, D=3, E=4) ──────────────────────────
            sect_l = str(row[1] or '').replace('\n', ' ').strip()
            if sect_l:
                if '주말' in sect_l:
                    cur = '주말'
                elif '평일' in sect_l:
                    cur = '평일'
            name_l  = str(row[3] or '').strip()
            sheet_l = '주말' if '/주말' in name_l else cur
            _register(name_l, row[4], sheet_l)

            # ── 오른쪽 테이블 (J=9, K=10, L=11, M=12) ─────────────────────
            name_r  = str(row[11] or '').strip() if len(row) > 11 else ''
            price_r = row[12] if len(row) > 12 else None
            sheet_r = '주말' if '/주말' in name_r else '평일'
            _register(name_r, price_r, sheet_r)

    except Exception:
        pass

    return pm


def find_price(course_name: str, sheet: str, pm: dict) -> int | None:
    """
    과정명 + 구분(평일/주말) → 수강료(원) or None.
    국비/계약 과정은 -1 반환 (무료 마커).
    """
    if not pm:
        return None
    if re.match(r'^\(국\)|^\(계\)', course_name):
        return -1  # 국비훈련

    def _lookup(name: str) -> int | None:
        entry = pm.get(name)
        if entry:
            return entry.get(sheet) or entry.get('평일') or next(iter(entry.values()), None)
        return None

    # 1) 정확 일치
    p = _lookup(course_name)
    if p:
        return p

    # 2) 공백 제거 후 일치
    name_ns = course_name.replace(' ', '')
    for key, entry in pm.items():
        if key.replace(' ', '') == name_ns:
            return entry.get(sheet) or entry.get('평일') or next(iter(entry.values()), None)

    # 3) 끝 숫자 제거: 마야3 → 마야, 블렌더1 → 블렌더
    name_base = re.sub(r'\d+$', '', course_name).strip()
    if name_base and name_base != course_name:
        p = _lookup(name_base)
        if p:
            return p

    # 4) 퍼지: 과정명이 price key의 시작이거나 price key가 과정명의 시작
    for key, entry in pm.items():
        k = key.replace(' ', '').lower()
        n = name_ns.lower()
        n_base = name_base.replace(' ', '').lower()
        if n.startswith(k) or k.startswith(n) or (n_base and k.startswith(n_base)):
            return entry.get(sheet) or entry.get('평일') or next(iter(entry.values()), None)

    return None


# ══════════════════════════════════════════════════════════════════════════════
# 엑셀 생성
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(student_name: str, cart: list[dict]) -> bytes:
    """★전체시간표 (양식).xlsx 기반으로 개인 시간표를 생성합니다."""

    if not TEMPLATE_PATH.exists():
        buf = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "템플릿 파일이 없습니다: " + str(TEMPLATE_PATH))
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    # ── 템플릿 로드 (서식·병합·테마 색상 그대로 유지) ────────────────────────
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active
    ws.title = f"{student_name} 시간표"

    ca = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ── 행1~2: 제목 (A1:K2 병합 유지) ──────────────────────────────────────
    ws['A1'].value     = f'SBS아카데미  {student_name}  개인 시간표'
    ws['A1'].alignment = ca

    # ── 행3: 연도 (A3:K3 병합 유지) ─────────────────────────────────────────
    ws['A3'].value     = f'{datetime.now().year}년'
    ws['A3'].alignment = ca

    wd_courses = [item for item in cart if item['sheet'] == '평일']
    we_courses = [item for item in cart if item['sheet'] == '주말']

    def _time_range(courses: list[dict]) -> str:
        """과정 목록에서 시간대 문자열 생성."""
        pairs: list[str] = []
        seen: set[str] = set()
        for c in courses:
            st = c.get('start_time', '')
            et = c.get('end_time', '')
            if st:
                label = f"{st}~{et}" if et else st
                if label not in seen:
                    seen.add(label)
                    pairs.append(label)
        return ' / '.join(pairs) if pairs else '-'

    def _course_cell_text(item: dict) -> str:
        """하나의 과정을 셀 텍스트로 변환 (줄바꿈 포함)."""
        lines = [item['name']]
        days = item.get('days', '')
        st = item.get('start_time', '')
        et = item.get('end_time', '')
        time_str = f"{st}{'~'+et if et else ''}" if st else ''
        if days and time_str:
            lines.append(f"{days}  {time_str}")
        elif time_str:
            lines.append(time_str)
        sd = item.get('start_date', '')
        ed = item.get('end_date', '')
        if sd or ed:
            lines.append(f"{sd}~{ed}" if sd and ed else sd or ed)
        return '\n'.join(lines)

    # ── 행4: 평일 헤더 (서식 유지, 값 없음) ─────────────────────────────────
    # ws['A4'] 값 '평일'은 템플릿 그대로 유지

    # ── 행5: 평일 과정 (A5=시간대, B5-K5=과정 수평 배치) ────────────────────
    ws['A5'].value     = _time_range(wd_courses) if wd_courses else '-'
    ws['A5'].alignment = ca
    for col_i, item in enumerate(wd_courses[:10], start=2):   # B=2 … K=11
        cell = ws.cell(5, col_i)
        cell.value     = _course_cell_text(item)
        cell.alignment = ca

    # ── 행6: 주말 헤더 (서식 유지, 값 없음) ─────────────────────────────────
    # ws['A6'] 값 '주말'은 템플릿 그대로 유지

    # ── 행7: 주말 과정 (A7=시간대, B7-K7=과정 수평 배치) ────────────────────
    ws['A7'].value     = _time_range(we_courses) if we_courses else '-'
    ws['A7'].alignment = ca
    for col_i, item in enumerate(we_courses[:10], start=2):
        cell = ws.cell(7, col_i)
        cell.value     = _course_cell_text(item)
        cell.alignment = ca

    # ── 행8~9: 구분선·비고 (템플릿 서식 유지) ────────────────────────────────
    # 변경 없음 — 템플릿 그대로 출력

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# Streamlit UI
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title='시간표 검색 | SBS아카데미',
    page_icon='🔍',
    layout='wide',
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700;900&display=swap');
html,body,[class*="css"]{font-family:'Noto Sans KR',sans-serif;}

/* ── 상단 헤더 ── */
.sbs-header{
  background:linear-gradient(120deg,#1e3a5f 0%,#1F4E79 60%,#2563a8 100%);
  padding:1.1rem 1.8rem;border-radius:12px;margin-bottom:.8rem;
  box-shadow:0 4px 20px rgba(31,78,121,.28);
}
.sbs-header h1{margin:0;color:#fff;font-size:1.35rem;font-weight:900;}
.sbs-header p{margin:.2rem 0 0;color:rgba(255,255,255,.78);font-size:.84rem;}

/* ── 요일 필터 ── */
.day-filter-wrap{
  display:flex;align-items:center;gap:.35rem;
  background:#f8fafc;border:1.5px solid #e2e8f0;
  border-radius:10px;padding:.45rem .8rem;margin-bottom:.6rem;
}
.day-filter-label{font-size:.78rem;font-weight:700;color:#64748b;white-space:nowrap;}

/* ── 결과 테이블 ── */
.tbl-wrap{
  border:1.5px solid #e2e8f0;border-radius:10px;
  overflow:hidden;margin-top:.3rem;
}
.tbl-head{
  display:grid;
  grid-template-columns: 2.6fr 1.5fr 1.4fr 1.1fr 1.1fr 0.8fr;
  background:#1e3a5f;padding:.45rem .6rem;
  font-size:.76rem;font-weight:800;color:#fff;gap:.3rem;
}
.tbl-row{
  display:grid;
  grid-template-columns: 2.6fr 1.5fr 1.4fr 1.1fr 1.1fr 0.8fr;
  padding:.42rem .6rem;gap:.3rem;align-items:center;
  border-bottom:1px solid #f1f5f9;
  transition:background .1s;
}
.tbl-row:last-child{border-bottom:none;}
.tbl-row:hover{background:#f0f7ff;}
.tbl-row.we-row{border-left:3px solid #059669;}
.tbl-row.wd-row{border-left:3px solid #2563a8;}

.c-name{font-size:.88rem;font-weight:800;color:#0f172a;line-height:1.3;}
.c-sheet{
  display:inline-block;border-radius:4px;padding:1px 5px;
  font-size:.65rem;font-weight:700;margin-left:.3rem;vertical-align:middle;
}
.c-sheet-wd{background:#dbeafe;color:#1d4ed8;}
.c-sheet-we{background:#dcfce7;color:#059669;}
.c-days{font-size:.84rem;font-weight:700;color:#1d4ed8;letter-spacing:.5px;}
.c-days-we{color:#059669;}
.c-time{font-size:.82rem;color:#374151;}
.c-date{font-size:.79rem;color:#475569;}
.c-room{font-size:.75rem;color:#92400e;background:#fef3c7;
  border-radius:4px;padding:1px 5px;display:inline-block;}

/* ── 요일 원형 뱃지 (소형) ── */
.day-dot{
  display:inline-flex;align-items:center;justify-content:center;
  width:1.3rem;height:1.3rem;border-radius:50%;
  font-size:.68rem;font-weight:800;margin:0 .05rem;
}
.day-dot.on-wd{background:#1F4E79;color:#fff;}
.day-dot.on-we{background:#059669;color:#fff;}
.day-dot.off{background:#e2e8f0;color:#94a3b8;}

/* ── 장바구니 카드 ── */
.cart-card{
  background:#f0f7ff;border:1.5px solid #bfdbfe;
  border-left:4px solid #2563a8;
  border-radius:0 9px 9px 0;padding:.5rem .85rem;margin-bottom:.35rem;
}
.cart-card.we{border-left-color:#059669;background:#f0fdf4;border-color:#bbf7d0;}
.cart-name{font-weight:800;color:#1e40af;font-size:.90rem;}
.cart-meta{font-size:.76rem;color:#64748b;line-height:1.7;margin-top:.12rem;}
.cart-period{font-weight:700;color:#1e3a5f;}
.cart-days{font-weight:700;color:#1d4ed8;font-size:.78rem;}

.price-tag{
  margin-left:auto;
  background:#fef9c3;color:#78350f;
  border:1.5px solid #fde68a;border-radius:6px;
  padding:.06rem .55rem;font-size:.82rem;font-weight:900;
  white-space:nowrap;letter-spacing:-.5px;
}
.price-tag.free{background:#dcfce7;color:#15803d;border-color:#bbf7d0;}
.price-tag.unknown{background:#f1f5f9;color:#94a3b8;border-color:#e2e8f0;font-weight:500;}

.cart-total{
  background:linear-gradient(90deg,#1e3a5f,#2563a8);
  color:#fff;border-radius:8px;padding:.5rem 1rem;
  display:flex;justify-content:space-between;align-items:center;
  margin-bottom:.5rem;
}
.cart-total-label{font-size:.80rem;opacity:.85;}
.cart-total-amount{font-size:1.05rem;font-weight:900;letter-spacing:-.5px;}

.stButton>button{border-radius:8px!important;}
.stDownloadButton>button{
  background:#1F4E79!important;color:#fff!important;
  border-radius:8px!important;font-weight:700!important;
}
</style>
""", unsafe_allow_html=True)

_hdr_col, _refresh_col = st.columns([5, 1])
with _hdr_col:
    st.markdown("""
<div class="sbs-header">
  <h1>🔍 SBS아카데미 강의 시간표 검색</h1>
  <p>과정명 검색 + 요일 필터로 원하는 강의를 빠르게 찾아보세요</p>
</div>
""", unsafe_allow_html=True)
with _refresh_col:
    st.markdown("<div style='margin-top:.6rem'></div>", unsafe_allow_html=True)
    if st.button('🔄 새로고침', help='시간표 파일이 업데이트됐을 때 캐시를 지우고 다시 불러옵니다.',
                 use_container_width=True):
        load_all_courses.clear()
        st.rerun()

# ── 데이터 체크 ───────────────────────────────────────────────────────────────
if not WEEKDAY_DIR.exists():
    st.error(f'⚠️ 시간표 폴더를 찾을 수 없습니다: {WEEKDAY_DIR}')
    st.stop()

all_courses = load_all_courses()
if not all_courses:
    st.error('⚠️ 강의시간표 파일을 읽을 수 없습니다.')
    st.stop()

# ── 로드 현황 표시 ─────────────────────────────────────────────────────────────
_wd_cnt  = sum(1 for c in all_courses if c.sheet == '평일')
_we_cnt  = sum(1 for c in all_courses if c.sheet == '주말')
_src_dir = WEEKDAY_DIR.name
_xls_cnt = len(list(WEEKDAY_DIR.glob('*.xls')))
st.markdown(
    f"<p style='font-size:.75rem;color:#64748b;margin:.0rem 0 .5rem'>"
    f"📂 데이터 소스: <b>{_src_dir}/</b> ({_xls_cnt}개 파일) &nbsp;·&nbsp; "
    f"평일 <b style='color:#1d4ed8'>{_wd_cnt}</b>개"
    + (f" &nbsp;·&nbsp; 주말 <b style='color:#059669'>{_we_cnt}</b>개" if _we_cnt else "")
    + "</p>",
    unsafe_allow_html=True,
)

price_map = load_price_map()

# ── 세션 상태 ─────────────────────────────────────────────────────────────────
if 'cart' not in st.session_state:
    st.session_state.cart = []
if 'sel_day' not in st.session_state:
    st.session_state.sel_day = '전체'


def add_to_cart(e: CourseEntry, price: int | None):
    key = (e.name, e.room, e.start_time, e.start_date)
    if not any(
        (c['name'], c['room'], c['start_time'], c['start_date']) == key
        for c in st.session_state.cart
    ):
        st.session_state.cart.append({
            'name':       e.name,
            'room':       e.room,
            'start_time': e.start_time,
            'end_time':   e.end_time,
            'instructor': e.instructor,
            'days':       e.days,
            'start_date': e.start_date,
            'end_date':   e.end_date,
            'sheet':      e.sheet,
            'source':     e.source,
            'price':      price,
        })


def _in_cart(e: CourseEntry) -> bool:
    key = (e.name, e.room, e.start_time, e.start_date)
    return any(
        (c['name'], c['room'], c['start_time'], c['start_date']) == key
        for c in st.session_state.cart
    )


def _filter_by_day(entries: list, day: str) -> list:
    """요일 필터 적용."""
    if not day or day == '전체':
        return entries
    if day == '주말':
        return [e for e in entries if _days_to_set(e.days) & {'토', '일'}]
    if day == '평일만':
        return [e for e in entries if _days_to_set(e.days) - {'토', '일'}]
    return [e for e in entries if day in _days_to_set(e.days)]


def _days_badge_row(days: str, is_we: bool = False) -> str:
    """요일 원형 뱃지 행 HTML (소형)."""
    active = _days_to_set(days)
    on_cls = 'on-we' if is_we else 'on-wd'
    parts = []
    for d in _DAY_ORDER:
        cls = on_cls if d in active else 'off'
        parts.append(f"<span class='day-dot {cls}'>{d}</span>")
    return ''.join(parts)


def _fmt_date(d: str) -> str:
    """YYYY-MM-DD → M/D"""
    try:
        dt = datetime.strptime(d, "%Y-%m-%d")
        return f"{dt.month}/{dt.day}"
    except Exception:
        return d


# ── 레이아웃 ──────────────────────────────────────────────────────────────────
col_left, col_right = st.columns([3, 2], gap='large')

# ══════════════════════════════════════════════════════════════════════════════
# 왼쪽: 검색 + 요일 필터 + 테이블
# ══════════════════════════════════════════════════════════════════════════════
with col_left:

    # ── 검색바 ──────────────────────────────────────────────────────────────
    keyword = st.text_input(
        '과정명 검색',
        placeholder='예) 에펙, 캐드, 마야, 포토샵, 컴활 …  (빈칸 = 전체 보기)',
        label_visibility='collapsed',
    )

    # ── 요일 필터 버튼 ────────────────────────────────────────────────────
    DAY_FILTERS = ['전체', '월', '화', '수', '목', '금', '토', '일', '주말', '평일만']
    st.markdown("<div style='margin-bottom:.2rem;font-size:.77rem;font-weight:700;color:#475569'>📅 요일 필터</div>", unsafe_allow_html=True)
    day_cols = st.columns(len(DAY_FILTERS))
    for col, label in zip(day_cols, DAY_FILTERS):
        is_active = st.session_state.sel_day == label
        btn_type = 'primary' if is_active else 'secondary'
        if col.button(label, key=f'dayf_{label}', type=btn_type, use_container_width=True):
            st.session_state.sel_day = label
            st.rerun()

    st.markdown("<div style='margin:.3rem 0'></div>", unsafe_allow_html=True)

    # ── 데이터 필터링 ──────────────────────────────────────────────────────
    kw = keyword.strip()
    base = search_courses(all_courses, kw) if kw else list(all_courses)
    results = _filter_by_day(base, st.session_state.sel_day)

    # ── 결과 통계 ──────────────────────────────────────────────────────────
    wd_cnt = sum(1 for e in results if e.sheet == '평일')
    we_cnt = sum(1 for e in results if e.sheet == '주말')
    stat_parts = [f"<b style='color:#1e3a5f'>{len(results)}개</b> 과정"]
    if wd_cnt:
        stat_parts.append(f"<span style='color:#1d4ed8'>평일 {wd_cnt}</span>")
    if we_cnt:
        stat_parts.append(f"<span style='color:#059669'>주말 {we_cnt}</span>")
    st.markdown(
        f"<p style='font-size:.80rem;color:#64748b;margin:.1rem 0 .4rem'>"
        + " &nbsp;·&nbsp; ".join(stat_parts) + "</p>",
        unsafe_allow_html=True,
    )

    if not results:
        st.markdown(
            "<div style='background:#fff7ed;border:1.5px solid #fed7aa;"
            "border-radius:9px;padding:1rem 1.2rem;color:#9a3412;font-size:.88rem'>"
            "검색 결과가 없습니다. 키워드나 요일 필터를 변경해 보세요.</div>",
            unsafe_allow_html=True,
        )
    else:
        # ── 테이블 헤더 ─────────────────────────────────────────────────
        st.markdown(
            "<div class='tbl-wrap'>"
            "<div class='tbl-head'>"
            "<span>과정명</span>"
            "<span>요일</span>"
            "<span>시간</span>"
            "<span>개강일</span>"
            "<span>종강일</span>"
            "<span>강의실</span>"
            "</div></div>",
            unsafe_allow_html=True,
        )

        # ── 테이블 행 (각 과정) ─────────────────────────────────────────
        for e in results:
            is_we     = e.sheet == '주말'
            row_cls   = 'we-row' if is_we else 'wd-row'
            days_cls  = 'c-days-we' if is_we else 'c-days'
            sheet_cls = 'c-sheet-we' if is_we else 'c-sheet-wd'
            _et       = getattr(e, 'end_time', '') or ''
            time_str  = f"{e.start_time}~{_et}" if _et else e.start_time
            price     = find_price(e.name, e.sheet, price_map)
            in_cart   = _in_cart(e)

            # HTML 정보 행 + Streamlit 버튼 열
            info_col, btn_col = st.columns([11, 1])
            with info_col:
                st.markdown(
                    f"<div class='tbl-row {row_cls}'>"
                    # 과정명
                    f"<span><span class='c-name'>{e.name}</span>"
                    f"<span class='c-sheet {sheet_cls}'>{e.sheet}</span></span>"
                    # 요일 (원형 뱃지)
                    f"<span>{_days_badge_row(e.days, is_we)}</span>"
                    # 시간
                    f"<span class='c-time'>⏰ {time_str}</span>"
                    # 개강일
                    f"<span class='c-date'>📅 {_fmt_date(e.start_date)}</span>"
                    # 종강일
                    f"<span class='c-date'>🏁 {_fmt_date(e.end_date)}</span>"
                    # 강의실
                    f"<span><span class='c-room'>📍{e.room}</span></span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            with btn_col:
                if in_cart:
                    st.markdown(
                        "<div style='text-align:center;color:#059669;font-size:1.1rem;"
                        "font-weight:900;padding-top:.3rem'>✓</div>",
                        unsafe_allow_html=True,
                    )
                else:
                    btn_key = f"add_{e.name}_{e.room}_{e.start_date}_{e.start_time}"
                    if st.button('＋', key=btn_key, use_container_width=True,
                                 help=f'{e.name} 장바구니에 추가'):
                        add_to_cart(e, price)
                        st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# 오른쪽: 장바구니
# ══════════════════════════════════════════════════════════════════════════════
with col_right:
    cart = st.session_state.cart
    st.markdown(
        f'### 🛒 선택 과목 &nbsp;'
        f'<span style="font-size:.85rem;color:#64748b;font-weight:500">({len(cart)}개)</span>',
        unsafe_allow_html=True,
    )

    if not cart:
        st.markdown(
            "<div style='background:#f8fafc;border:2px dashed #cbd5e1;"
            "border-radius:10px;padding:2rem;text-align:center;color:#94a3b8'>"
            "<div style='font-size:2rem;margin-bottom:.5rem'>🛒</div>"
            "<div style='font-weight:600;margin-bottom:.3rem'>선택한 과목이 없어요</div>"
            "<small>왼쪽 검색 결과에서 <b>＋</b>를 누르세요</small>"
            '</div>',
            unsafe_allow_html=True,
        )
    else:
        for i, item in enumerate(cart):
            is_we     = item['sheet'] == '주말'
            card_cls  = 'we' if is_we else ''
            sheet_cls = 'c-sheet-we' if is_we else 'c-sheet-wd'
            days_cls  = 'c-days-we' if is_we else 'c-days'

            try:
                sd = datetime.strptime(item['start_date'], "%Y-%m-%d")
                ed = datetime.strptime(item['end_date'],   "%Y-%m-%d")
                period_str = f"{sd.year}년 {sd.month}.{sd.day} ~ {ed.month}.{ed.day}"
            except Exception:
                period_str = f"{item['start_date']} ~ {item['end_date']}"

            p = item.get('price')
            if p == -1:
                price_html = "<span class='price-tag free' style='font-size:.72rem'>국비 무료</span>"
            elif isinstance(p, int) and p > 0:
                price_html = f"<span class='price-tag' style='font-size:.74rem'>{p:,}원</span>"
            else:
                price_html = "<span class='price-tag unknown' style='font-size:.72rem'>수강료 문의</span>"

            col_info, col_del = st.columns([6, 1])
            with col_info:
                st.markdown(
                    f"<div class='cart-card {card_cls}'>"
                    f"<div style='display:flex;align-items:center;gap:.35rem;margin-bottom:.18rem'>"
                    f"<span class='cart-name'>{item['name']}</span>"
                    f"<span class='c-sheet {sheet_cls}'>{item['sheet']}</span>"
                    f"{price_html}"
                    f"</div>"
                    f"<div class='cart-meta'>"
                    f"<span class='{days_cls}'>"
                    f"{_days_badge_row(item.get('days',''), is_we)}"
                    f"</span> "
                    f"⏰ {item['start_time']}{'~'+item['end_time'] if item.get('end_time') else ''}"
                    f" &nbsp;·&nbsp; 📍 {item['room']}<br>"
                    f"<span class='cart-period'>🗓 {period_str}</span>"
                    f"</div></div>",
                    unsafe_allow_html=True,
                )
            with col_del:
                st.markdown("<div style='margin-top:.3rem'></div>", unsafe_allow_html=True)
                if st.button('✕', key=f'del_{i}', help='제거'):
                    st.session_state.cart.pop(i)
                    st.rerun()

        # ── 수강료 합계 ──────────────────────────────────────────────────────
        known_prices  = [item['price'] for item in cart
                         if isinstance(item.get('price'), int) and item['price'] > 0]
        gov_count     = sum(1 for item in cart if item.get('price') == -1)
        unknown_count = len(cart) - len(known_prices) - gov_count

        if known_prices or gov_count:
            total_str = f"{sum(known_prices):,}원"
            if gov_count:
                total_str += f" + 국비 {gov_count}개"
            if unknown_count:
                total_str += f" + 문의 {unknown_count}개"
            note = " (국비 제외)" if gov_count else (" (문의 제외)" if unknown_count else "")
            st.markdown(
                f"<div class='cart-total'>"
                f"<span class='cart-total-label'>💰 예상 수강료 합계{note}</span>"
                f"<span class='cart-total-amount'>{total_str}</span>"
                f"</div>",
                unsafe_allow_html=True,
            )

        st.markdown("<div style='margin-top:.3rem'></div>", unsafe_allow_html=True)
        if st.button('🗑️ 전체 비우기', use_container_width=True):
            st.session_state.cart = []
            st.rerun()

        st.markdown('---')

        # ── 엑셀 생성 ────────────────────────────────────────────────────────
        st.markdown('#### 📥 개인 시간표 생성')
        sname = st.text_input(
            '수강생 이름',
            placeholder='홍길동',
            key='sname',
        )
        if st.button('📊 엑셀 시간표 만들기', type='primary', use_container_width=True):
            if not sname.strip():
                st.warning('수강생 이름을 입력해 주세요.')
            else:
                with st.spinner('생성 중...'):
                    xlsx = build_excel(sname.strip(), cart)
                safe  = re.sub(r'[\\/:*?"<>|]', '_', sname.strip())
                fname = f"SBS_{safe}_시간표_{datetime.now().strftime('%Y%m%d')}.xlsx"
                st.success('✅ 생성 완료!')
                st.download_button(
                    '⬇️ 다운로드',
                    data=xlsx,
                    file_name=fname,
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True,
                )

# ── 푸터 ──────────────────────────────────────────────────────────────────────
st.markdown('---')
st.markdown(
    "<p style='text-align:center;color:#9ca3af;font-size:.78rem'>"
    'SBS아카데미 대전지점 · 강의 시간표 검색 시스템</p>',
    unsafe_allow_html=True,
)
